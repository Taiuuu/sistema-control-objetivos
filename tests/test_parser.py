from pathlib import Path

import openpyxl

from services.importador.parser import leer_pasadas_crudas
from services.importador.reporte import _construir_pasadas_normalizadas
from services.importador.duplicados import detectar_duplicados_internos
from services.importador.modelos import PasadaNormalizada
from services.importador.matcher import matchear_objetivo
from services.importador.normalizador import normalizar_hora
from services.importador.parser import leer_hojas_de_datos
from datetime import date, time


_ENCABEZADOS = ["NO", "OBJETIVO", "TURNO", "MOVIL", "HORA", "SUPERVISOR"]


def test_lee_pasada_de_bloque_siguiente_si_el_primero_esta_vacio(tmp_path: Path):
    ruta = tmp_path / "recorridos.xlsx"
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "1-8 (D)"

    for indice, encabezado in enumerate(_ENCABEZADOS, start=1):
        hoja.cell(row=1, column=indice, value=encabezado)
        hoja.cell(row=1, column=indice + 7, value=encabezado)

    hoja.cell(row=2, column=8, value=1)
    hoja.cell(row=2, column=9, value="Objetivo B")
    hoja.cell(row=2, column=10, value="D")
    hoja.cell(row=2, column=11, value="M-2")
    hoja.cell(row=2, column=12, value="08:00")
    hoja.cell(row=2, column=13, value="Supervisor")
    libro.save(ruta)

    pasadas = leer_pasadas_crudas(str(ruta), "1-8 (D)")

    no_vacias = [pasada for pasada in pasadas if not pasada.esta_vacia()]
    assert len(no_vacias) == 1
    assert no_vacias[0].bloque_tabla == 2
    assert no_vacias[0].objetivo == "Objetivo B"


def test_lee_tres_bloques_normaliza_objetivo_y_corta_observaciones(tmp_path: Path):
    ruta = tmp_path / "recorridos.xlsx"
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "17-8 (D)"

    for inicio in (1, 8, 15):
        for indice, encabezado in enumerate(_ENCABEZADOS):
            hoja.cell(row=2, column=inicio + indice, value=encabezado)

    hoja.cell(row=3, column=1, value=1)
    hoja.cell(row=3, column=2, value="OBRA VILLA DE MAYO ")
    hoja.cell(row=3, column=3, value="D")
    hoja.cell(row=3, column=4, value="M-1")
    hoja.cell(row=3, column=5, value="09;11")
    hoja.cell(row=3, column=6, value="Supervisor")
    hoja.cell(row=3, column=10, value="D")
    hoja.cell(row=3, column=11, value="M-2")
    hoja.cell(row=3, column=12, value="10:00")
    hoja.cell(row=3, column=13, value="Supervisor")
    hoja.cell(row=3, column=18, value="D")
    hoja.cell(row=3, column=19, value="M-3")
    hoja.cell(row=3, column=20, value="11:00")
    hoja.cell(row=3, column=21, value="Supervisor")
    hoja.cell(row=4, column=1, value="OBSERVACIONES: otras labores")
    hoja.cell(row=5, column=1, value="texto libre")
    libro.save(ruta)

    pasadas = [
        pasada
        for pasada in leer_pasadas_crudas(str(ruta), hoja.title)
        if not pasada.esta_vacia()
    ]

    assert [pasada.bloque_tabla for pasada in pasadas] == [1, 2, 3]
    assert pasadas[0].objetivo == "OBRA VILLA DE MAYO"


def test_excel_agosto_conserva_todas_las_pasadas_reales():
    ruta = Path(__file__).parents[1] / "CONTROL RECORRIDOS AGOSTO 2026 (2).xlsx"

    pasadas, problemas, hojas, detectadas = _construir_pasadas_normalizadas(
        str(ruta), 2026
    )

    assert hojas == 62
    assert detectadas == 1430
    assert len(pasadas) == 1422
    assert sum("no cargó la hora" in problema.descripcion for problema in problemas) == 8


def test_duplicados_no_mezcla_objetivos_sin_id():
    base = dict(
        hoja="1-8 (D)",
        fila_excel=3,
        bloque_tabla=1,
        fecha_operativa=date(2026, 8, 1),
        fecha_calendario=date(2026, 8, 1),
        turno="D",
        hora=time(10, 0),
        supervisor_nombre="Supervisor",
        supervisor_id=None,
        objetivo_id=None,
    )
    primera = PasadaNormalizada(**base, objetivo_nombre="OBJETIVO A")
    segunda = PasadaNormalizada(**{**base, "fila_excel": 4}, objetivo_nombre="OBJETIVO B")

    finales, descartadas = detectar_duplicados_internos([primera, segunda])

    assert len(finales) == 2
    assert descartadas == []


def _guardar_libro_con_bloques(ruta: Path, nombre: str, filas: dict[int, dict]):
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = nombre
    for inicio in (1, 8, 15):
        for indice, encabezado in enumerate(_ENCABEZADOS):
            hoja.cell(2, inicio + indice, encabezado)
    for fila, bloques in filas.items():
        for inicio, valores in bloques.items():
            for indice, valor in enumerate(valores):
                hoja.cell(fila, inicio + indice, valor)
    libro.save(ruta)


def test_pasa_bloque_uno_y_dos(tmp_path: Path):
    ruta = tmp_path / "dos_bloques.xlsx"
    _guardar_libro_con_bloques(
        ruta,
        "1-8 (D)",
        {3: {1: [1, "OBJETIVO", "D", "M-1", "10:00", "SUP"] , 8: [1, "OBJETIVO", "D", "M-2", "11:00", "SUP"]}},
    )
    pasadas = [p for p in leer_pasadas_crudas(str(ruta), "1-8 (D)" ) if not p.esta_vacia()]
    assert [p.bloque_tabla for p in pasadas] == [1, 2]


def test_objetivo_sin_pasada_no_se_cuenta(tmp_path: Path):
    ruta = tmp_path / "sin_pasada.xlsx"
    _guardar_libro_con_bloques(
        ruta,
        "1-8 (D)",
        {3: {1: [1, "OBJETIVO", None, None, None, None], 8: [1, "OBJETIVO", None, None, None, None], 15: [1, "OBJETIVO", None, None, None, None]}},
    )
    assert not [p for p in leer_pasadas_crudas(str(ruta), "1-8 (D)") if not p.esta_vacia()]


def test_pasada_parcial_se_conserva_sin_hora(tmp_path: Path):
    ruta = tmp_path / "parcial.xlsx"
    _guardar_libro_con_bloques(
        ruta,
        "4-8 (N)",
        {3: {1: [1, "OBJETIVO", "N", "M-1", None, "SUP"]}},
    )
    pasadas = [p for p in leer_pasadas_crudas(str(ruta), "4-8 (N)") if not p.esta_vacia()]
    assert len(pasadas) == 1
    assert pasadas[0].hora is None


def test_hoja_desplazada_y_hojas_no_datos_se_filtran(tmp_path: Path):
    ruta = tmp_path / "hojas.xlsx"
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "17-8 (D)"
    for indice, encabezado in enumerate(_ENCABEZADOS):
        hoja.cell(2, 2 + indice, encabezado)
    hoja.cell(3, 2, 1); hoja.cell(3, 3, "OBJETIVO"); hoja.cell(3, 4, "D"); hoja.cell(3, 5, "M"); hoja.cell(3, 6, "09;11"); hoja.cell(3, 7, "SUP")
    listado = libro.create_sheet("Listado")
    listado["A1"] = "TURNO"
    libro.create_sheet("Hoja 18")
    libro.save(ruta)
    assert leer_hojas_de_datos(str(ruta)) == ["17-8 (D)"]
    pasadas = [p for p in leer_pasadas_crudas(str(ruta), "17-8 (D)") if not p.esta_vacia()]
    assert pasadas[0].hora == "09;11"


def test_hora_con_punto_y_coma_y_objetivo_con_espacios():
    resultado = normalizar_hora("09;11")
    assert resultado.hora == time(9, 11)
    assert matchear_objetivo("OBRA VILLA DE MAYO ", [{"id": 1, "nombre": "OBRA VILLA DE MAYO"}]).tipo == "exacto"
