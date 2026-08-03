import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from services.importador_universal import ImportadorUniversal


def test_parsear_nombre_sheet_acepta_barra():
    importador = ImportadorUniversal()

    resultado = importador._parsear_nombre_sheet("27/5 (D)")

    assert resultado == (date.today().replace(month=5, day=27), "diurno")


def test_parsear_hoja_control_recorridos_limpia_texto():
    importador = ImportadorUniversal()
    wb = Workbook()
    ws = wb.active
    ws.title = "27-5 (D)"

    ws["A1"] = "Objetivo"
    ws["B1"] = "Supervisor"
    ws["C1"] = "Hora"
    ws["D1"] = "Veces"
    ws["E1"] = "Notas"

    ws["A2"] = "Revisar\nplaca"
    ws["B2"] = "Juan   Perez"
    ws["C2"] = "08:30"
    ws["D2"] = "1"
    ws["E2"] = "sin observaciones"

    registros = importador._parsear_hoja_control_recorridos(ws, date(2026, 5, 27), "diurno")

    assert len(registros) == 1
    assert registros[0].objetivo == "Revisar placa"
    assert registros[0].supervisor == "Juan Perez"
    assert registros[0].notas == "sin observaciones"


def test_es_encabezado_detecta_aliases_de_notas_y_rechaza_substrings():
    importador = ImportadorUniversal()

    assert importador._es_encabezado("Objetivo") is True
    assert importador._es_encabezado("SUPERVISOR") is True
    assert importador._es_encabezado("Hora") is True
    assert importador._es_encabezado("Turno") is True
    assert importador._es_encabezado("Notas") is True
    assert importador._es_encabezado("Observaciones") is True
    assert importador._es_encabezado("Observación") is True
    assert importador._es_encabezado("Horacio") is False
    assert importador._es_encabezado("Objetivos cumplidos") is False
    assert importador._es_encabezado("Horas Hombre") is False
    assert importador._es_encabezado("Supervisión") is False


def test_buscar_encabezados_control_rechaza_falsos_positivos_por_substrings():
    importador = ImportadorUniversal()
    wb = Workbook()
    ws = wb.active
    ws.title = "27-5 (D)"

    ws["A1"] = "Objetivos cumplidos"
    ws["B1"] = "Horas Hombre"
    ws["C1"] = "Supervisor Horacio"
    ws["A2"] = "Objetivo A"
    ws["B2"] = "08:30"
    ws["C2"] = "Juan Perez"

    header_row, columnas = importador._buscar_encabezados_control(ws)

    assert header_row is None
    assert columnas == {}


def test_listar_sheet_options_y_parsear_control_recorridos_comparten_criterio_para_legacy_control():
    importador = ImportadorUniversal()
    wb = Workbook()
    ws = wb.active
    ws.title = "CONTROL DE RECORRIDOS"
    ws["B2"] = "Objetivo A"
    ws["E2"] = "08:30"
    ws["F2"] = "Juan"

    opciones = importador._listar_sheet_options(wb)
    resultado = importador._parsear_control_recorridos(wb)

    assert opciones[0]["title"] == "CONTROL DE RECORRIDOS"
    assert opciones[0]["fecha"] == date.today().isoformat()
    assert opciones[0]["turno"] is None
    assert len(resultado["registros"]) == 1


def test_metricas_del_parser_registran_hojas_filas_y_registros():
    importador = ImportadorUniversal()
    wb = Workbook()
    ws = wb.active
    ws.title = "27-5 (D)"
    ws["A1"] = "Objetivo"
    ws["B1"] = "Supervisor"
    ws["C1"] = "Hora"
    ws["A2"] = "Objetivo A"
    ws["B2"] = "Juan"
    ws["C2"] = "08:30"
    ws["A3"] = "Objetivo B"
    ws["B3"] = "Ana"
    ws["C3"] = "Cerrado"
    ws["A4"] = "Objetivo C"
    ws["B4"] = ""
    ws["C4"] = ""
    ws.row_dimensions[5].hidden = True
    ws["A5"] = "Objetivo oculto"
    ws["B5"] = "Pedro"
    ws["C5"] = "09:00"

    ws_vacia = wb.create_sheet("Vacía")
    ws_vacia["A1"] = ""

    importador._parsear_control_recorridos(wb)

    metrics = importador._metrics
    assert metrics.workbook.total_hojas == 2
    assert metrics.workbook.hojas_analizadas == 2
    assert metrics.workbook.hojas_validas == 1
    assert metrics.workbook.hojas_descartadas == 1
    assert metrics.filas.total_filas_recorridas >= 5
    assert metrics.filas.filas_vacias >= 1
    assert metrics.filas.filas_ocultas >= 1
    assert metrics.filas.filas_encabezado >= 1
    assert metrics.extraccion.registros_creados >= 2


def test_invalida_cache_del_importador_al_invalidar_objetivos_y_supervisores(monkeypatch):
    import services.cache as cache_module
    import services.importador_universal as importador_module

    importador = ImportadorUniversal()

    monkeypatch.setattr(importador_module.gestor_db, 'ejecutar', lambda query, params=None: [])

    importador.reload_cache()
    assert importador._cache_inicializado is True

    cache_module.invalidar_objetivos()
    assert importador._cache_inicializado is False

    cache_module.invalidar_supervisores()
    assert importador._cache_inicializado is False


def test_invalidate_cache_y_reload_cache_recargan_datos(monkeypatch):
    import services.importador_universal as importador_module

    importador = ImportadorUniversal()

    respuestas = [
        [
            {'id': 1, 'nombre': 'Objetivo viejo'},
            {'id': 2, 'nombre': 'Objetivo nuevo'},
        ],
        [
            {'id': 10, 'nombre': 'Supervisor viejo'},
            {'id': 11, 'nombre': 'Supervisor nuevo'},
        ],
    ]

    def fake_ejecutar(query, params=None):
        if 'FROM objetivos' in query:
            return respuestas[0]
        if 'FROM supervisores' in query:
            return respuestas[1]
        return []

    monkeypatch.setattr(importador_module.gestor_db, 'ejecutar', fake_ejecutar)

    importador.reload_cache()

    assert importador._obtener_objetivo_id('Objetivo viejo') == 1
    assert importador._obtener_supervisor_id('Supervisor nuevo') == 11

    respuestas[0] = [{'id': 5, 'nombre': 'Objetivo actualizado'}]
    respuestas[1] = [{'id': 20, 'nombre': 'Supervisor actualizado'}]

    importador.invalidate_cache()
    importador.reload_cache()

    assert importador._obtener_objetivo_id('Objetivo actualizado') == 5
    assert importador._obtener_supervisor_id('Supervisor actualizado') == 20
    assert importador._obtener_objetivo_id('Objetivo viejo') is None
    assert importador._obtener_supervisor_id('Supervisor viejo') is None

def test_inicializar_caches_no_marca_valido_si_falla_la_carga(monkeypatch):
    import services.importador_universal as importador_module

    importador = ImportadorUniversal()

    def ejecutar_falla(query, params=None):
        raise RuntimeError("DB caída")

    monkeypatch.setattr(importador_module.gestor_db, 'ejecutar', ejecutar_falla)

    importador._inicializar_caches()

    assert importador._cache_inicializado is False
    assert importador._cache_objetivos == {}

def test_parsear_con_encabezados_control_respecta_turno_del_excel():
    importador = ImportadorUniversal()
    wb = Workbook()
    ws = wb.active
    ws.title = "27-5 (D)"

    ws["A1"] = "Objetivo"
    ws["B1"] = "Supervisor"
    ws["C1"] = "Hora"
    ws["D1"] = "Turno"

    ws["A2"] = "Pase prueba"
    ws["B2"] = "Juan Perez"
    ws["C2"] = "03:15"
    ws["D2"] = "Nocturno"

    header_row, columnas = importador._buscar_encabezados_control(ws)
    registros = importador._parsear_con_encabezados_control(
        ws,
        date(2026, 5, 27),
        "diurno",
        columnas,
        header_row,
    )

    assert len(registros) == 1
    assert registros[0].turno == "nocturno"


def test_normalizar_hora_y_fecha_acepta_separadores_repetidos():
    importador = ImportadorUniversal()

    fecha, hora = importador._normalizar_hora_y_fecha(
        "16::14",
        date(2026, 5, 28),
    )

    assert fecha == date(2026, 5, 28)
    assert hora == "16:14"


def test_normalizar_hora_y_fecha_rechaza_cerrado_como_hora_invalida():
    importador = ImportadorUniversal()

    try:
        importador._normalizar_hora_y_fecha("Cerrado ", date(2026, 5, 28))
        assert False, "Se esperaba ValueError para hora inválida"
    except ValueError as exc:
        assert "Hora vacía o inválida" in str(exc)


def test_parsear_control_recorridos_legacy_ignora_fila_global_y_detecta_datos_en_fila_dos():
    importador = ImportadorUniversal()
    wb = Workbook()
    ws = wb.active
    ws.title = "18-5 (N)"

    ws["A1"] = "CONTROL DE RECORRIDOS"
    ws["A2"] = 11
    ws["B2"] = "CENTRO INTEGRAL DE LA MUJER"
    ws["C2"] = "NOCHE"
    ws["D2"] = "L-200"
    ws["E2"] = "00:28"
    ws["F2"] = "LUCIANO, DANIEL ACACIO"

    ws["H2"] = 11
    ws["I2"] = "CENTRO INTEGRAL DE LA MUJER"
    ws["J2"] = "NOCHE"
    ws["K2"] = "HILUX"
    ws["L2"] = "07:18"
    ws["M2"] = "GONZALEZ, MAXIMILIANO"

    registros = importador._parsear_control_recorridos_legacy(ws, date(2026, 5, 18), "nocturno")

    assert len(registros) == 2
    assert registros[0].objetivo == "CENTRO INTEGRAL DE LA MUJER"
    assert registros[1].objetivo == "CENTRO INTEGRAL DE LA MUJER"


def test_parsers_control_recorridos_comparten_logica_de_registros():
    importador = ImportadorUniversal()

    wb_encabezados = Workbook()
    ws_encabezados = wb_encabezados.active
    ws_encabezados.title = "27-5 (D)"
    ws_encabezados["A1"] = "Objetivo"
    ws_encabezados["B1"] = "Supervisor"
    ws_encabezados["C1"] = "Hora"
    ws_encabezados["A2"] = "Objetivo A"
    ws_encabezados["B2"] = "Juan"
    ws_encabezados["C2"] = "08:30"
    ws_encabezados["A3"] = "Objetivo B"
    ws_encabezados["B3"] = ""
    ws_encabezados["C3"] = ""

    header_row, columnas = importador._buscar_encabezados_control(ws_encabezados)
    registros_encabezados = importador._parsear_con_encabezados_control(
        ws_encabezados,
        date(2026, 5, 27),
        "diurno",
        columnas,
        header_row,
    )

    wb_legacy = Workbook()
    ws_legacy = wb_legacy.active
    ws_legacy.title = "27-5 (D)"
    ws_legacy["B2"] = "Objetivo A"
    ws_legacy["C2"] = ""
    ws_legacy["E2"] = "08:30"
    ws_legacy["F2"] = "Juan"
    ws_legacy["D2"] = "nota"
    ws_legacy["B3"] = "Objetivo B"
    ws_legacy["C3"] = ""
    ws_legacy["E3"] = ""
    ws_legacy["F3"] = ""

    registros_legacy = importador._parsear_control_recorridos_legacy(
        ws_legacy,
        date(2026, 5, 27),
        "diurno",
    )

    assert len(registros_encabezados) == 2
    assert len(registros_legacy) == 2

    for idx in range(2):
        datos_encabezados = registros_encabezados[idx].to_dict()
        datos_legacy = registros_legacy[idx].to_dict()
        for clave in ('fecha', 'hora', 'turno', 'supervisor', 'objetivo', 'fuente', 'sheet_title'):
            assert datos_encabezados[clave] == datos_legacy[clave], (idx, clave, datos_encabezados, datos_legacy)

    assert registros_encabezados[1].hora == ""
    assert registros_encabezados[1].supervisor == ""


def test_crear_pasada_offline_calcula_fecha_operativa(monkeypatch):
    import services.sync_manager as sync_module

    captured = {}

    class DummyProvider:
        def crear_pasada(self, pasada):
            captured["fecha_operativa"] = pasada.fecha_operativa
            return True

    def fake_agregar_cambio_pendiente(self, tipo, operacion, datos):
        captured["cambio"] = {
            "tipo": tipo,
            "operacion": operacion,
            "datos": datos,
        }

    monkeypatch.setattr(sync_module, "get_data_provider", lambda: DummyProvider())
    monkeypatch.setattr(sync_module.SyncManager, "_guardar_cambios_pendientes", lambda self: None)
    monkeypatch.setattr(sync_module.SyncManager, "_cargar_cambios_pendientes", lambda self: None)

    manager = sync_module.SyncManager.__new__(sync_module.SyncManager)
    manager.cambios_pendientes = []
    manager.agregar_cambio_pendiente = fake_agregar_cambio_pendiente.__get__(manager, sync_module.SyncManager)

    resultado = manager.crear_pasada_offline(
        fecha="2026-05-26",
        hora="03:00",
        turno="nocturno",
        supervisor_id=7,
        objetivo_id=9,
        notas="prueba",
    )

    assert resultado is True
    assert captured["fecha_operativa"] == "2026-05-25"
    assert captured["cambio"]["datos"]["fecha_operativa"] == "2026-05-25"


def test_procesar_registros_control_recorridos_acepta_horas_fuera_de_rango(monkeypatch):
    import services.sync_manager as sync_module
    from services.importador_universal import RegistroImportacion

    importador = ImportadorUniversal()
    importador.sync_manager = sync_module.get_sync_manager()

    class DummyProvider:
        def crear_pasada(self, pasada):
            return True

    class DummySyncManager:
        def agregar_cambio_pendiente(self, *args, **kwargs):
            return None

        def crear_pasada_offline(self, *args, **kwargs):
            return True

    monkeypatch.setattr(sync_module, "get_data_provider", lambda: DummyProvider())
    monkeypatch.setattr(sync_module.SyncManager, "_guardar_cambios_pendientes", lambda self: None)
    monkeypatch.setattr(sync_module.SyncManager, "_cargar_cambios_pendientes", lambda self: None)
    monkeypatch.setattr(importador, "_obtener_supervisor_id", lambda nombre: 7)
    monkeypatch.setattr(importador, "_obtener_objetivo_id", lambda nombre: 9)
    monkeypatch.setattr(importador, "sync_manager", DummySyncManager())
    monkeypatch.setattr("services.gestor_turnos.GestorTurnos.calcular_fecha_operativa", lambda *args, **kwargs: (_ for _ in ()).throw(ValueError("Hora fuera del rango")))

    registros = [
        RegistroImportacion(
            fecha="2026-05-27",
            hora="03:11",
            turno="diurno",
            supervisor="Supervisor test",
            objetivo="Objetivo test",
            notas="prueba",
            fuente="excel",
            sheet_title="27-5 (D)",
        )
    ]

    resultado = importador._procesar_registros(registros)

    assert resultado.registros_validos == 1
    assert resultado.registros_errores == 0
    assert resultado.exitoso is True


def test_no_agrega_objetivos_faltantes(monkeypatch):
    import services.sync_manager as sync_module
    from services.importador_universal import RegistroImportacion

    importador = ImportadorUniversal()
    importador.sync_manager = sync_module.get_sync_manager()

    class DummyProvider:
        def crear_pasada(self, pasada):
            return True

    class DummySyncManager:
        def agregar_cambio_pendiente(self, *args, **kwargs):
            return None

        def crear_pasada_offline(self, *args, **kwargs):
            return True

    monkeypatch.setattr(sync_module, "get_data_provider", lambda: DummyProvider())
    monkeypatch.setattr(sync_module.SyncManager, "_guardar_cambios_pendientes", lambda self: None)
    monkeypatch.setattr(sync_module.SyncManager, "_cargar_cambios_pendientes", lambda self: None)
    monkeypatch.setattr(importador, "_obtener_supervisor_id", lambda nombre: 7)
    monkeypatch.setattr(importador, "_obtener_objetivo_id", lambda nombre: None)
    monkeypatch.setattr(importador, "sync_manager", DummySyncManager())

    registros = [
        RegistroImportacion(
            fecha="2026-05-27",
            hora="08:30",
            turno="diurno",
            supervisor="Supervisor test",
            objetivo="OBJETIVO INEXISTENTE",
            notas="prueba",
            fuente="excel",
            sheet_title="27-5 (D)",
        )
    ]

    resultado = importador._procesar_registros(registros)

    assert resultado.registros_validos == 0
    assert resultado.registros_errores == 1
    assert "Objetivo no encontrado" in resultado.errores[0]
    assert resultado.exitoso is False


def test_filtrar_registros_por_rango_respecta_limites_fecha_y_turno():
    from services.importador_universal import RegistroImportacion

    importador = ImportadorUniversal()
    registros = [
        RegistroImportacion(fecha="2026-06-01", hora="08:00", turno="diurno", supervisor="S", objetivo="O", fuente="excel", sheet_title="01-6 (D)"),
        RegistroImportacion(fecha="2026-06-11", hora="00:30", turno="nocturno", supervisor="S", objetivo="O", fuente="excel", sheet_title="11-6 (N)"),
        RegistroImportacion(fecha="2026-06-12", hora="08:00", turno="diurno", supervisor="S", objetivo="O", fuente="excel", sheet_title="12-6 (D)"),
        RegistroImportacion(fecha="2026-06-13", hora="00:30", turno="nocturno", supervisor="S", objetivo="O", fuente="excel", sheet_title="13-6 (N)"),
        RegistroImportacion(fecha="2026-06-13", hora="08:00", turno="diurno", supervisor="S", objetivo="O", fuente="excel", sheet_title="13-6 (D)"),
        RegistroImportacion(fecha="2026-06-30", hora="00:30", turno="nocturno", supervisor="S", objetivo="O", fuente="excel", sheet_title="30-6 (N)"),
    ]

    filtrados = importador.filtrar_registros_por_rango(
        registros,
        rango_desde=(date(2026, 6, 11), "nocturno"),
        rango_hasta=(date(2026, 6, 13), "diurno"),
    )

    assert [r.fecha for r in filtrados] == ["2026-06-11", "2026-06-12", "2026-06-13", "2026-06-13"]
    assert [r.turno for r in filtrados] == ["nocturno", "diurno", "nocturno", "diurno"]
