"""Verificacion independiente de pasadas de CONTROL_RECORRIDOS.

Uso en Windows:
    python scripts\verificar_pasadas_excel.py "CONTROL RECORRIDOS AGOSTO 2026 (2).xlsx"

Con pipeline completo y una base SQLite:
    python scripts\verificar_pasadas_excel.py archivo.xlsx --db seguridad.db
"""

from __future__ import annotations

import argparse
import logging
import re
import sqlite3
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from services.importador import parser
from services.importador.reporte import analizar_excel

PATRON_HOJA = re.compile(r"^\d{1,2}-\d{1,2} \((D|N)\)$")
ENCABEZADOS = ("NO", "OBJETIVO", "TURNO", "MOVIL", "HORA", "SUPERVISOR")
logger = logging.getLogger("verificar_pasadas_excel")


def presente(valor: Any) -> bool:
    return valor is not None and (not isinstance(valor, str) or bool(valor.strip()))


def encontrar_bloques(ws) -> tuple[int, list[int]]:
    for fila in range(1, min(ws.max_row, 15) + 1):
        inicios = []
        for columna in range(1, ws.max_column - len(ENCABEZADOS) + 2):
            valores = tuple(
                str(ws.cell(fila, columna + desplazamiento).value).strip().upper()
                if ws.cell(fila, columna + desplazamiento).value is not None
                else ""
                for desplazamiento in range(len(ENCABEZADOS))
            )
            if valores == ENCABEZADOS:
                inicios.append(columna)
        if inicios:
            return fila, inicios
    raise ValueError(f"{ws.title}: no se encontraron encabezados completos")


def contar_manual(ws) -> tuple[int, list[tuple[int, int, str]]]:
    fila_encabezado, inicios = encontrar_bloques(ws)
    conteo = 0
    detalles = []
    for fila in range(fila_encabezado + 1, ws.max_row + 1):
        primera = ws.cell(fila, 1).value
        if isinstance(primera, str) and primera.strip().upper().startswith("OBSERVACIONES"):
            break
        for bloque, inicio in enumerate(inicios[:3], start=1):
            valores_pase = [ws.cell(fila, inicio + desplazamiento).value for desplazamiento in range(2, 6)]
            if any(presente(valor) for valor in valores_pase):
                objetivo = ws.cell(fila, inicio + 1).value
                conteo += 1
                detalles.append((fila, bloque, str(objetivo).strip() if objetivo is not None else ""))
    return conteo, detalles


def main() -> int:
    argumentos = argparse.ArgumentParser()
    argumentos.add_argument("archivo", type=Path)
    argumentos.add_argument("--db", type=Path)
    argumentos.add_argument("--log-level", default="WARNING")
    opciones = argumentos.parse_args()
    logging.basicConfig(level=getattr(logging, opciones.log_level.upper()), format="%(levelname)s %(message)s")

    workbook = openpyxl.load_workbook(opciones.archivo, data_only=True)
    manual_por_hoja = {}
    discrepancias = []
    print("hoja;manual;parser;diferencia")
    for nombre in workbook.sheetnames:
        if not PATRON_HOJA.fullmatch(nombre.strip()):
            continue
        manual, detalles = contar_manual(workbook[nombre])
        crudas = [p for p in parser.leer_pasadas_crudas(str(opciones.archivo), nombre) if not p.esta_vacia()]
        parser_count = len(crudas)
        manual_por_hoja[nombre] = manual
        diferencia = parser_count - manual
        print(f"{nombre};{manual};{parser_count};{diferencia}")
        if diferencia:
            parser_keys = {(p.fila_excel, p.bloque_tabla) for p in crudas}
            manual_keys = {(fila, bloque) for fila, bloque, _ in detalles}
            discrepancias.append((nombre, manual_keys - parser_keys, parser_keys - manual_keys))

    manual_total = sum(manual_por_hoja.values())
    parser_total = sum(len([p for p in parser.leer_pasadas_crudas(str(opciones.archivo), nombre) if not p.esta_vacia()]) for nombre in manual_por_hoja)
    print(f"TOTAL;{manual_total};{parser_total};{parser_total - manual_total}")
    if discrepancias:
        for nombre, faltantes, extras in discrepancias:
            print(f"DISCREPANCIA {nombre}: faltantes={sorted(faltantes)} extras={sorted(extras)}")
        return 1

    if opciones.db:
        with sqlite3.connect(opciones.db) as conexion:
            resultado = analizar_excel(str(opciones.archivo), 2026, conexion)
        pipeline_por_hoja = defaultdict(int)
        for pasada in resultado.pasadas:
            pipeline_por_hoja[pasada.hoja] += 1
        print("\nhoja;parser;pipeline_normalizadas;diferencia")
        for nombre in manual_por_hoja:
            crudas = [p for p in parser.leer_pasadas_crudas(str(opciones.archivo), nombre) if not p.esta_vacia()]
            parser_count = len(crudas)
            pipeline_count = pipeline_por_hoja[nombre]
            print(f"{nombre};{parser_count};{pipeline_count};{pipeline_count - parser_count}")
            if pipeline_count != parser_count:
                for pasada in crudas:
                    if pasada.hora is None:
                        print(
                            "PIPELINE_DISCREPANCIA "
                            f"hoja={pasada.hoja} fila={pasada.fila_excel} "
                            f"bloque={pasada.bloque_tabla} "
                            f"objetivo={pasada.objetivo!r} "
                            "motivo=hora ausente, queda como advertencia "
                            "y no se convierte en registro persistible"
                        )
        print(f"PIPELINE;{parser_total};{len(resultado.pasadas)};{len(resultado.pasadas) - parser_total}")
        print(f"PIPELINE_DETECTADAS;{resultado.pasadas_detectadas};sin_hora={resultado.pasadas_sin_hora};duplicadas={resultado.pasadas_duplicadas}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
