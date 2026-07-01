#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de diagnóstico para entender la estructura del Excel CONTROL_RECORRIDOS
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))

def diagnosticar_excel(ruta_archivo: str):
    """Analiza la estructura del Excel"""
    
    wb = load_workbook(ruta_archivo, data_only=True)
    
    print("\n" + "="*80)
    print(f"DIAGNÓSTICO: {Path(ruta_archivo).name}")
    print("="*80)
    
    # Listar hojas
    print(f"\nTotal de hojas: {len(wb.worksheets)}")
    for idx, ws in enumerate(wb.worksheets, 1):
        print(f"  {idx}. {ws.title}")
    
    # Analizar la primera hoja o la del 17-8
    hoja_objetivo = None
    for ws in wb.worksheets:
        if '17-8' in ws.title or '17' in ws.title:
            hoja_objetivo = ws
            break
    
    if hoja_objetivo is None:
        hoja_objetivo = wb.worksheets[0]
    
    print(f"\n\nAnalizando hoja: '{hoja_objetivo.title}'")
    print("="*80)
    
    # Detectar encabezados
    print("\nPrimeras 5 filas (completas):")
    print("-" * 200)
    
    for row_idx in range(1, min(6, hoja_objetivo.max_row + 1)):
        fila = []
        for col_idx in range(1, min(21, hoja_objetivo.max_column + 1)):  # A-T = 1-20
            valor = hoja_objetivo.cell(row=row_idx, column=col_idx).value
            valor_str = str(valor).strip() if valor else ""
            fila.append(f"{valor_str:15}")
        
        col_labels = "".join([f"{chr(64+i):15}" for i in range(1, 21)])
        if row_idx == 1:
            print(col_labels)
            print("-" * 200)
        
        print(f"Fila {row_idx}: {''.join(fila)}")
    
    # Buscar bloques (grupos separados por columnas vacías)
    print("\n\nDetectando bloques de datos:")
    print("-" * 80)
    
    # Analizar estructura de columnas
    for row_idx in range(2, min(10, hoja_objetivo.max_row + 1)):
        fila = []
        for col_idx in range(1, 21):
            valor = hoja_objetivo.cell(row=row_idx, column=col_idx).value
            fila.append(valor)
        
        # Encontrar bloques: grupos de columnas con datos
        bloques = []
        en_bloque = False
        inicio_bloque = None
        
        for col_idx, valor in enumerate(fila, start=1):
            tiene_dato = valor is not None and str(valor).strip() != ""
            
            if tiene_dato and not en_bloque:
                en_bloque = True
                inicio_bloque = col_idx
            elif not tiene_dato and en_bloque:
                en_bloque = False
                bloques.append((inicio_bloque, col_idx - 1))
        
        if en_bloque:
            bloques.append((inicio_bloque, len(fila)))
        
        if bloques:
            print(f"\nFila {row_idx} - Bloques detectados:")
            for inicio, fin in bloques:
                cols_letras = "".join([chr(64 + i) for i in range(inicio, fin + 1)])
                datos = [str(fila[i-1])[:10] for i in range(inicio, fin + 1)]
                print(f"  Cols {cols_letras} ({inicio}-{fin}): {datos}")
    
    # Información sobre max_row y max_column
    print(f"\n\nMetadatos de la hoja:")
    print(f"  max_row: {hoja_objetivo.max_row}")
    print(f"  max_column: {hoja_objetivo.max_column}")
    print(f"  Letra de última columna: {chr(64 + hoja_objetivo.max_column)}")
    
    # Verificar si hay datos en filas posteriores
    print(f"\n\nVerificando filas 40-50:")
    tiene_datos = False
    for row_idx in range(40, min(51, hoja_objetivo.max_row + 1)):
        fila_vacia = True
        for col_idx in range(1, 21):
            valor = hoja_objetivo.cell(row=row_idx, column=col_idx).value
            if valor and str(valor).strip():
                fila_vacia = False
                break
        
        if not fila_vacia:
            tiene_datos = True
            print(f"  Fila {row_idx}: tiene datos")
    
    if not tiene_datos:
        print(f"  (sin datos en filas 40-50)")
    
    print("\n" + "="*80)


def main():
    # Buscar archivos Excel en el directorio
    carpeta = Path(__file__).parent
    archivos_excel = list(carpeta.glob("*.xlsx")) + list(carpeta.glob("*.xls"))
    
    if not archivos_excel:
        print("No se encontraron archivos Excel en el directorio")
        sys.exit(1)
    
    print(f"\nEncontrados {len(archivos_excel)} archivos Excel")
    
    for archivo in archivos_excel:
        if "CONTROL" in archivo.name.upper():
            print(f"\n✓ Analizando: {archivo.name}")
            try:
                diagnosticar_excel(str(archivo))
            except Exception as e:
                print(f"✗ Error: {e}")
                import traceback
                traceback.print_exc()


if __name__ == "__main__":
    main()
