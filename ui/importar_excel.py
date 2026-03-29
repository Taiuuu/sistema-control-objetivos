# =============================================================================
# VESP Organizations - Sistema de Control de Objetivos
# Pantalla para importar datos desde Excel existente
# Soporta formato con bloque diurno y nocturno en columnas separadas
# =============================================================================

import sqlite3
import datetime
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QFileDialog, QMessageBox, QTextEdit,
    QComboBox, QSpinBox
)
from PyQt6.QtCore import Qt
from database.db import DB_PATH
from openpyxl import load_workbook


# =============================================================================
# FUNCIONES DE IMPORTACIÓN
# =============================================================================

def _obtener_o_crear_objetivo(cursor, nombre: str) -> int:
    """Busca un objetivo por nombre o lo crea si no existe. Retorna su ID."""
    cursor.execute("SELECT id FROM objetivos WHERE nombre = ?", (nombre,))
    resultado = cursor.fetchone()
    if resultado:
        return resultado[0]

    hoy = datetime.datetime.now().strftime("%Y-%m-%d")
    cursor.execute("""
        INSERT INTO objetivos (nombre, fecha_inicio, dias_semana)
        VALUES (?, ?, ?)
    """, (nombre, hoy, "1,2,3,4,5,6,7"))
    return cursor.lastrowid


def _obtener_o_crear_supervisor(cursor, nombre: str) -> int | None:
    """Busca un supervisor por nombre o lo crea si no existe. Retorna su ID."""
    if not nombre or str(nombre).strip() in ("", "None", "none"):
        return None

    nombre = str(nombre).strip()
    cursor.execute("SELECT id FROM supervisores WHERE nombre = ?", (nombre,))
    resultado = cursor.fetchone()
    if resultado:
        return resultado[0]

    cursor.execute("INSERT INTO supervisores (nombre) VALUES (?)", (nombre,))
    return cursor.lastrowid


def _normalizar_fecha(fecha_raw) -> str:
    """Convierte distintos formatos de fecha a yyyy-MM-dd."""
    if isinstance(fecha_raw, datetime.datetime):
        return fecha_raw.strftime("%Y-%m-%d")
    if isinstance(fecha_raw, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.datetime.strptime(fecha_raw, fmt).strftime("%Y-%m-%d")
            except Exception:
                pass
    return datetime.datetime.now().strftime("%Y-%m-%d")


def importar_bloque(cursor, ws, turno: str, col_supervisor: int,
                    col_objetivo: int, col_veces: int,
                    col_fecha: int | None, fila_inicio: int) -> tuple:
    """
    Importa un bloque de pasadas (diurno o nocturno) desde el Excel.
    Retorna (pasadas_importadas, errores).
    """
    pasadas = 0
    errores = []

    for fila_num, fila in enumerate(ws.iter_rows(min_row=fila_inicio, values_only=True), fila_inicio):
        try:
            if not fila or all(v is None for v in fila):
                continue

            objetivo_nombre = str(fila[col_objetivo - 1]).strip() if fila[col_objetivo - 1] else None
            supervisor_nombre = str(fila[col_supervisor - 1]).strip() if fila[col_supervisor - 1] else ""
            veces_raw = fila[col_veces - 1]
            fecha_raw = fila[col_fecha - 1] if col_fecha else None

            if not objetivo_nombre or objetivo_nombre.lower() in ("none", "objetivo", ""):
                continue

            try:
                veces = int(float(str(veces_raw))) if veces_raw else 0
            except Exception:
                veces = 0

            if veces == 0:
                continue

            fecha = _normalizar_fecha(fecha_raw)
            objetivo_id = _obtener_o_crear_objetivo(cursor, objetivo_nombre)
            supervisor_id = _obtener_o_crear_supervisor(cursor, supervisor_nombre)

            for _ in range(veces):
                cursor.execute("""
                    INSERT INTO pasadas (fecha, hora, turno, objetivo_id, supervisor_id)
                    VALUES (?, ?, ?, ?, ?)
                """, (fecha, "No especificado", turno, objetivo_id, supervisor_id))

            pasadas += veces

        except Exception as e:
            errores.append(f"Fila {fila_num} ({turno}): {e}")

    return pasadas, errores


# =============================================================================
# PANTALLA DE IMPORTACIÓN
# =============================================================================

class ImportarExcel(QWidget):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Importar desde Excel")
        self.setGeometry(200, 200, 650, 600)
        self.ruta_archivo = None

        layout = QVBoxLayout()

        titulo = QLabel("Importar datos desde Excel")
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo.setStyleSheet("font-size: 16px; font-weight: bold; color: #4CAF50;")
        layout.addWidget(titulo)

        desc = QLabel(
            "El Excel tiene dos bloques: uno para turno diurno y otro para nocturno.\n"
            "Indicá en qué columna está cada dato para cada bloque.\n"
            "Las columnas se numeran desde 1 (A=1, B=2, C=3...)"
        )
        desc.setAlignment(Qt.AlignmentFlag.AlignCenter)
        desc.setStyleSheet("color: #888; font-size: 11px;")
        layout.addWidget(desc)

        layout.addSpacing(8)

        # Selector de archivo
        fila_archivo = QHBoxLayout()
        self.label_archivo = QLabel("Ningún archivo seleccionado")
        self.label_archivo.setStyleSheet("color: #888;")
        boton_archivo = QPushButton("Seleccionar Excel")
        boton_archivo.clicked.connect(self._seleccionar_archivo)
        fila_archivo.addWidget(self.label_archivo)
        fila_archivo.addWidget(boton_archivo)
        layout.addLayout(fila_archivo)

        layout.addSpacing(8)

        # Fila de inicio
        fila_inicio_layout = QHBoxLayout()
        fila_inicio_layout.addWidget(QLabel("Fila donde empiezan los datos:"))
        self.fila_inicio = QSpinBox()
        self.fila_inicio.setMinimum(1)
        self.fila_inicio.setMaximum(20)
        self.fila_inicio.setValue(3)
        fila_inicio_layout.addWidget(self.fila_inicio)
        fila_inicio_layout.addStretch()
        layout.addLayout(fila_inicio_layout)

        layout.addSpacing(8)

        # Bloque diurno
        label_diurno = QLabel("☀ Bloque DIURNO")
        label_diurno.setStyleSheet("font-size: 13px; font-weight: bold; color: #FFD700;")
        layout.addWidget(label_diurno)

        grid_diurno = QHBoxLayout()
        self.diurno_supervisor = self._combo_col("Supervisor:", 3, grid_diurno)
        self.diurno_objetivo = self._combo_col("Objetivo:", 4, grid_diurno)
        self.diurno_veces = self._combo_col("Veces:", 6, grid_diurno)
        self.diurno_fecha = self._combo_col_opcional("Fecha:", grid_diurno)
        layout.addLayout(grid_diurno)

        layout.addSpacing(8)

        # Bloque nocturno
        label_nocturno = QLabel("🌙 Bloque NOCTURNO")
        label_nocturno.setStyleSheet("font-size: 13px; font-weight: bold; color: #4FC3F7;")
        layout.addWidget(label_nocturno)

        grid_nocturno = QHBoxLayout()
        self.nocturno_supervisor = self._combo_col("Supervisor:", 8, grid_nocturno)
        self.nocturno_objetivo = self._combo_col("Objetivo:", 9, grid_nocturno)
        self.nocturno_veces = self._combo_col("Veces:", 10, grid_nocturno)
        self.nocturno_fecha = self._combo_col_opcional("Fecha:", grid_nocturno)
        layout.addLayout(grid_nocturno)

        layout.addSpacing(10)

        boton_importar = QPushButton("Importar datos")
        boton_importar.setFixedHeight(40)
        boton_importar.clicked.connect(self._importar)
        layout.addWidget(boton_importar)

        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setFixedHeight(100)
        self.log.setStyleSheet("font-size: 11px;")
        layout.addWidget(self.log)

        self.setLayout(layout)

    def _combo_col(self, label: str, default: int, layout) -> QComboBox:
        """Crea un combo de selección de columna con label."""
        col_layout = QVBoxLayout()
        col_layout.addWidget(QLabel(label))
        combo = QComboBox()
        combo.addItems([str(i) for i in range(1, 21)])
        combo.setCurrentIndex(default - 1)
        col_layout.addWidget(combo)
        layout.addLayout(col_layout)
        return combo

    def _combo_col_opcional(self, label: str, layout) -> QComboBox:
        """Crea un combo de selección de columna opcional."""
        col_layout = QVBoxLayout()
        col_layout.addWidget(QLabel(label))
        combo = QComboBox()
        combo.addItems(["No hay"] + [str(i) for i in range(1, 21)])
        col_layout.addWidget(combo)
        layout.addLayout(col_layout)
        return combo

    def _seleccionar_archivo(self) -> None:
        """Abre el diálogo para seleccionar el archivo Excel."""
        ruta, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar Excel", "",
            "Excel (*.xlsx *.xls)"
        )
        if ruta:
            self.ruta_archivo = ruta
            self.label_archivo.setText(ruta.split("/")[-1].split("\\")[-1])
            self.label_archivo.setStyleSheet("color: #4CAF50;")

    def _importar(self) -> None:
        """Ejecuta la importación de ambos bloques."""
        if not self.ruta_archivo:
            QMessageBox.warning(self, "Error", "Seleccioná un archivo Excel primero.")
            return

        try:
            wb = load_workbook(self.ruta_archivo, data_only=True)
            ws = wb.active

            conexion = sqlite3.connect(DB_PATH)
            cursor = conexion.cursor()

            self.log.clear()
            self.log.append("Importando bloque diurno...")

            diurno_fecha = None if self.diurno_fecha.currentText() == "No hay" else int(self.diurno_fecha.currentText())
            nocturno_fecha = None if self.nocturno_fecha.currentText() == "No hay" else int(self.nocturno_fecha.currentText())

            pasadas_dia, errores_dia = importar_bloque(
                cursor, ws, "diurno",
                col_supervisor=int(self.diurno_supervisor.currentText()),
                col_objetivo=int(self.diurno_objetivo.currentText()),
                col_veces=int(self.diurno_veces.currentText()),
                col_fecha=diurno_fecha,
                fila_inicio=self.fila_inicio.value()
            )

            self.log.append(f"✓ Diurno: {pasadas_dia} pasadas importadas.")

            self.log.append("Importando bloque nocturno...")

            pasadas_noche, errores_noche = importar_bloque(
                cursor, ws, "nocturno",
                col_supervisor=int(self.nocturno_supervisor.currentText()),
                col_objetivo=int(self.nocturno_objetivo.currentText()),
                col_veces=int(self.nocturno_veces.currentText()),
                col_fecha=nocturno_fecha,
                fila_inicio=self.fila_inicio.value()
            )

            self.log.append(f"✓ Nocturno: {pasadas_noche} pasadas importadas.")

            conexion.commit()
            conexion.close()

            total = pasadas_dia + pasadas_noche
            errores = errores_dia + errores_noche

            if errores:
                self.log.append(f"⚠ {len(errores)} errores encontrados:")
                for e in errores[:5]:
                    self.log.append(f"  {e}")

            from services.logger import registrar_accion
            from services.sesion import get_usuario_id
            registrar_accion(get_usuario_id(), f"Importó Excel: {total} pasadas ({pasadas_dia} diurnas, {pasadas_noche} nocturnas)")

            QMessageBox.information(
                self, "Listo",
                f"Importación completada.\n"
                f"Diurno: {pasadas_dia} pasadas\n"
                f"Nocturno: {pasadas_noche} pasadas\n"
                f"Total: {total} pasadas"
            )

        except Exception as e:
            self.log.append(f"✗ Error: {e}")
            QMessageBox.critical(self, "Error", f"No se pudo importar: {e}")