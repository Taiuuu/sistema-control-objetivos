# =============================================================================
# VESP Organizations - Sistema de Control de Objetivos
# Reporte detallado por objetivo: día a día con cobertura diurna/nocturna
# =============================================================================

import sqlite3
import datetime
import calendar
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QTableWidget, QTableWidgetItem,
    QComboBox, QFileDialog, QMessageBox
)
from PyQt6.QtGui import QColor
from services.background_task import run_background_task
from services.exportar import exportar_excel, exportar_pdf
from database.db import DB_PATH


# =============================================================================
# CÁLCULO DEL REPORTE POR OBJETIVO
# =============================================================================

def calcular_reporte_objetivo(anio: int, mes: int, objetivo_id: int) -> dict:
    """
    Retorna un dict con:
      - nombre: str
      - dias: list de dicts {fecha, dia_semana, diurno, nocturno, estado}
      - resumen: dict {dias_esperados, dias_con_dia, dias_con_noche, dias_sin_control, porcentaje}
    """
    conexion = sqlite3.connect(DB_PATH)
    cursor = conexion.cursor()

    cursor.execute(
        "SELECT nombre, fecha_inicio, fecha_fin, dias_semana FROM objetivos WHERE id = ?",
        (objetivo_id,)
    )
    row = cursor.fetchone()
    if not row:
        conexion.close()
        raise ValueError(f"No se encontró el objetivo con id {objetivo_id}")

    nombre, inicio, fin, dias_str = row
    dias_semana = [int(d) for d in dias_str.split(",")]
    total_dias = calendar.monthrange(anio, mes)[1]

    NOMBRES_DIA = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

    dias = []
    dias_esperados = 0
    dias_con_dia = 0
    dias_con_noche = 0
    dias_sin_control = 0

    for dia in range(1, total_dias + 1):
        fecha = f"{anio}-{mes:02d}-{dia:02d}"
        fecha_dt = datetime.datetime.strptime(fecha, "%Y-%m-%d")

        if inicio and fecha < inicio:
            continue
        if fin and fecha > fin:
            continue
        if fecha_dt.isoweekday() not in dias_semana:
            continue

        dias_esperados += 1
        nombre_dia = NOMBRES_DIA[fecha_dt.weekday()]

        cursor.execute("""
            SELECT COUNT(*) FROM pasadas
            WHERE fecha = ? AND objetivo_id = ? AND turno = 'diurno'
        """, (fecha, objetivo_id))
        tuvo_dia = cursor.fetchone()[0] > 0

        cursor.execute("""
            SELECT COUNT(*) FROM pasadas
            WHERE fecha = ? AND objetivo_id = ? AND turno = 'nocturno'
        """, (fecha, objetivo_id))
        tuvo_noche = cursor.fetchone()[0] > 0

        if tuvo_dia:
            dias_con_dia += 1
        if tuvo_noche:
            dias_con_noche += 1
        if not tuvo_dia and not tuvo_noche:
            dias_sin_control += 1

        if tuvo_dia and tuvo_noche:
            estado = "Completo"
        elif tuvo_dia:
            estado = "Solo diurno"
        elif tuvo_noche:
            estado = "Solo nocturno"
        else:
            estado = "Sin control"

        dias.append({
            "fecha": f"{dia:02d}/{mes:02d}/{anio}",
            "dia_semana": nombre_dia,
            "diurno": "✔" if tuvo_dia else "✘",
            "nocturno": "✔" if tuvo_noche else "✘",
            "estado": estado,
        })

    porcentaje = 0.0
    if dias_esperados > 0:
        dias_controlados = dias_esperados - dias_sin_control
        porcentaje = (dias_controlados / dias_esperados) * 100

    conexion.close()

    return {
        "nombre": nombre,
        "dias": dias,
        "resumen": {
            "dias_esperados": dias_esperados,
            "dias_con_dia": dias_con_dia,
            "dias_con_noche": dias_con_noche,
            "dias_sin_control": dias_sin_control,
            "porcentaje": porcentaje,
        }
    }


def cargar_objetivos_del_mes(anio: int, mes: int) -> list:
    """
    Devuelve lista de (id, nombre) de objetivos activos en ese mes.
    """
    conexion = sqlite3.connect(DB_PATH)
    cursor = conexion.cursor()
    primer_dia = f"{anio}-{mes:02d}-01"
    ultimo_dia = f"{anio}-{mes:02d}-{calendar.monthrange(anio, mes)[1]:02d}"

    cursor.execute("""
        SELECT id, nombre FROM objetivos
        WHERE (fecha_fin IS NULL OR fecha_fin >= ?)
          AND (fecha_inicio IS NULL OR fecha_inicio <= ?)
        ORDER BY nombre
    """, (primer_dia, ultimo_dia))
    objetivos = cursor.fetchall()
    conexion.close()
    return objetivos


# =============================================================================
# EXPORTACIÓN
# =============================================================================

def _exportar_excel_objetivo(anio: int, mes: int, objetivo_id: int, ruta: str) -> None:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    datos = calcular_reporte_objetivo(anio, mes, objetivo_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Título
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Reporte: {datos['nombre']} — {mes:02d}/{anio}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Encabezados
    encabezados = ["Fecha", "Día", "Diurno", "Nocturno", "Estado"]
    COLORES_HEADER = "2B4F8C"
    for col, enc in enumerate(encabezados, start=1):
        cell = ws.cell(row=3, column=col, value=enc)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORES_HEADER)
        cell.alignment = Alignment(horizontal="center")

    # Colores por estado
    COLOR_COMPLETO   = "90EE90"
    COLOR_PARCIAL    = "FFD700"
    COLOR_SIN        = "FF6B6B"

    def color_fila(estado):
        if estado == "Completo":
            return COLOR_COMPLETO
        elif estado in ("Solo diurno", "Solo nocturno"):
            return COLOR_PARCIAL
        return COLOR_SIN

    # Filas
    for fila_idx, d in enumerate(datos["dias"], start=4):
        valores = [d["fecha"], d["dia_semana"], d["diurno"], d["nocturno"], d["estado"]]
        fill = PatternFill("solid", fgColor=color_fila(d["estado"]))
        for col, val in enumerate(valores, start=1):
            cell = ws.cell(row=fila_idx, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    # Resumen
    r = datos["resumen"]
    fila_res = len(datos["dias"]) + 5
    ws.cell(row=fila_res,     column=1, value="Días esperados:").font    = Font(bold=True)
    ws.cell(row=fila_res,     column=2, value=r["dias_esperados"])
    ws.cell(row=fila_res + 1, column=1, value="Días c/ diurno:").font   = Font(bold=True)
    ws.cell(row=fila_res + 1, column=2, value=r["dias_con_dia"])
    ws.cell(row=fila_res + 2, column=1, value="Días c/ nocturno:").font = Font(bold=True)
    ws.cell(row=fila_res + 2, column=2, value=r["dias_con_noche"])
    ws.cell(row=fila_res + 3, column=1, value="Días sin control:").font = Font(bold=True)
    ws.cell(row=fila_res + 3, column=2, value=r["dias_sin_control"])
    ws.cell(row=fila_res + 4, column=1, value="Cumplimiento:").font     = Font(bold=True)
    ws.cell(row=fila_res + 4, column=2, value=f"{r['porcentaje']:.1f}%")

    # Ancho columnas
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16

    wb.save(ruta)


def _exportar_pdf_objetivo(anio: int, mes: int, objetivo_id: int, ruta: str) -> None:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    datos = calcular_reporte_objetivo(anio, mes, objetivo_id)
    doc = SimpleDocTemplate(ruta, pagesize=A4)
    styles = getSampleStyleSheet()
    elementos = []

    # Título
    elementos.append(Paragraph(
        f"<b>Reporte: {datos['nombre']}</b>",
        styles["Title"]
    ))
    elementos.append(Paragraph(
        f"Período: {mes:02d}/{anio}",
        styles["Normal"]
    ))
    elementos.append(Spacer(1, 12))

    # Tabla de días
    encabezados = ["Fecha", "Día", "Diurno", "Nocturno", "Estado"]
    filas = [encabezados]
    for d in datos["dias"]:
        filas.append([d["fecha"], d["dia_semana"], d["diurno"], d["nocturno"], d["estado"]])

    tabla = Table(filas, colWidths=[80, 40, 55, 65, 90])

    COLOR_HEADER   = colors.HexColor("#2B4F8C")
    COLOR_COMPLETO = colors.HexColor("#90EE90")
    COLOR_PARCIAL  = colors.HexColor("#FFD700")
    COLOR_SIN      = colors.HexColor("#FF6B6B")

    estilo_base = [
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_HEADER),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
    ]

    def color_estado(estado):
        if estado == "Completo":
            return COLOR_COMPLETO
        elif estado in ("Solo diurno", "Solo nocturno"):
            return COLOR_PARCIAL
        return COLOR_SIN

    for i, d in enumerate(datos["dias"], start=1):
        estilo_base.append(
            ("BACKGROUND", (0, i), (-1, i), color_estado(d["estado"]))
        )

    tabla.setStyle(TableStyle(estilo_base))
    elementos.append(tabla)
    elementos.append(Spacer(1, 16))

    # Resumen
    r = datos["resumen"]
    resumen_texto = (
        f"<b>Días esperados:</b> {r['dias_esperados']}  |  "
        f"<b>Con diurno:</b> {r['dias_con_dia']}  |  "
        f"<b>Con nocturno:</b> {r['dias_con_noche']}  |  "
        f"<b>Sin control:</b> {r['dias_sin_control']}  |  "
        f"<b>Cumplimiento:</b> {r['porcentaje']:.1f}%"
    )
    elementos.append(Paragraph(resumen_texto, styles["Normal"]))

    doc.build(elementos)


# =============================================================================
# PANTALLA
# =============================================================================

class ReporteObjetivo(QWidget):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Reporte por objetivo")
        self.setGeometry(200, 200, 750, 500)
        self._objetivos = []   # lista de (id, nombre)
        self._datos = None     # último reporte generado

        layout = QVBoxLayout()

        # --- Fila de selección ---
        fila = QHBoxLayout()

        self.selector_mes = QComboBox()
        meses = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
                 "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
        for m in meses:
            self.selector_mes.addItem(m)
        self.selector_mes.setCurrentIndex(datetime.datetime.now().month - 1)
        self.selector_mes.currentIndexChanged.connect(self._actualizar_objetivos)

        self.selector_anio = QComboBox()
        anio_actual = datetime.datetime.now().year
        for a in range(anio_actual - 2, anio_actual + 2):
            self.selector_anio.addItem(str(a))
        self.selector_anio.setCurrentText(str(anio_actual))
        self.selector_anio.currentIndexChanged.connect(self._actualizar_objetivos)

        self.selector_objetivo = QComboBox()
        self.selector_objetivo.setMinimumWidth(200)

        self.boton_generar = QPushButton("Generar reporte")
        self.boton_generar.clicked.connect(self._generar)

        self.boton_excel = QPushButton("Exportar Excel")
        self.boton_excel.clicked.connect(self._exportar_excel)
        self.boton_excel.setEnabled(False)

        self.boton_pdf = QPushButton("Exportar PDF")
        self.boton_pdf.clicked.connect(self._exportar_pdf)
        self.boton_pdf.setEnabled(False)

        fila.addWidget(QLabel("Mes:"))
        fila.addWidget(self.selector_mes)
        fila.addWidget(QLabel("Año:"))
        fila.addWidget(self.selector_anio)
        fila.addWidget(QLabel("Objetivo:"))
        fila.addWidget(self.selector_objetivo)
        fila.addWidget(self.boton_generar)
        fila.addWidget(self.boton_excel)
        fila.addWidget(self.boton_pdf)
        layout.addLayout(fila)

        self.estado_label = QLabel("Seleccioná un objetivo y generá el reporte.")
        layout.addWidget(self.estado_label)

        # --- Tabla ---
        self.tabla = QTableWidget()
        self.tabla.setColumnCount(5)
        self.tabla.setHorizontalHeaderLabels(
            ["Fecha", "Día", "Diurno", "Nocturno", "Estado"]
        )
        self.tabla.setColumnWidth(0, 110)
        self.tabla.setColumnWidth(1, 60)
        self.tabla.setColumnWidth(2, 80)
        self.tabla.setColumnWidth(3, 90)
        self.tabla.setColumnWidth(4, 120)
        self.tabla.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.tabla)

        # --- Resumen ---
        self.resumen_label = QLabel("")
        layout.addWidget(self.resumen_label)

        self.setLayout(layout)
        self._actualizar_objetivos()

    # ------------------------------------------------------------------
    def _get_mes_anio(self):
        return (
            self.selector_mes.currentIndex() + 1,
            int(self.selector_anio.currentText())
        )

    def _actualizar_objetivos(self) -> None:
        mes, anio = self._get_mes_anio()
        self._objetivos = cargar_objetivos_del_mes(anio, mes)
        self.selector_objetivo.clear()
        for obj_id, nombre in self._objetivos:
            self.selector_objetivo.addItem(nombre, userData=obj_id)
        self.boton_excel.setEnabled(False)
        self.boton_pdf.setEnabled(False)
        self._datos = None

    def _set_controls_enabled(self, enabled: bool) -> None:
        self.selector_mes.setEnabled(enabled)
        self.selector_anio.setEnabled(enabled)
        self.selector_objetivo.setEnabled(enabled)
        self.boton_generar.setEnabled(enabled)

    def _get_objetivo_id(self) -> int | None:
        idx = self.selector_objetivo.currentIndex()
        if idx < 0:
            return None
        return self.selector_objetivo.itemData(idx)

    # ------------------------------------------------------------------
    def _generar(self) -> None:
        mes, anio = self._get_mes_anio()
        objetivo_id = self._get_objetivo_id()
        if objetivo_id is None:
            QMessageBox.warning(self, "Atención", "No hay objetivos disponibles para ese mes.")
            return

        self._set_controls_enabled(False)
        self.boton_excel.setEnabled(False)
        self.boton_pdf.setEnabled(False)
        self.estado_label.setText("Generando reporte...")

        task = run_background_task(calcular_reporte_objetivo, anio, mes, objetivo_id)
        task.signals.finished.connect(self._on_reporte_generado)
        task.signals.error.connect(self._on_error)

    def _on_reporte_generado(self, datos: dict) -> None:
        self._datos = datos
        dias = datos["dias"]
        r = datos["resumen"]

        COLOR_COMPLETO  = QColor("#90EE90")
        COLOR_PARCIAL   = QColor("#FFD700")
        COLOR_SIN       = QColor("#FF6B6B")
        COLOR_TEXTO     = QColor("#000000")

        def color_estado(estado):
            if estado == "Completo":
                return COLOR_COMPLETO
            elif estado in ("Solo diurno", "Solo nocturno"):
                return COLOR_PARCIAL
            return COLOR_SIN

        self.tabla.setUpdatesEnabled(False)
        self.tabla.clearContents()
        self.tabla.setRowCount(len(dias))

        for i, d in enumerate(dias):
            valores = [d["fecha"], d["dia_semana"], d["diurno"], d["nocturno"], d["estado"]]
            color = color_estado(d["estado"])
            for col, val in enumerate(valores):
                item = QTableWidgetItem(val)
                item.setBackground(color)
                item.setForeground(COLOR_TEXTO)
                self.tabla.setItem(i, col, item)

        self.tabla.setUpdatesEnabled(True)

        self.resumen_label.setText(
            f"Días esperados: {r['dias_esperados']}  |  "
            f"Con diurno: {r['dias_con_dia']}  |  "
            f"Con nocturno: {r['dias_con_noche']}  |  "
            f"Sin control: {r['dias_sin_control']}  |  "
            f"Cumplimiento: {r['porcentaje']:.1f}%"
        )
        self.estado_label.setText(f"Reporte generado: {datos['nombre']}")
        self._set_controls_enabled(True)
        self.boton_excel.setEnabled(True)
        self.boton_pdf.setEnabled(True)

    def _on_error(self, mensaje: str) -> None:
        QMessageBox.critical(self, "Error", mensaje)
        self.estado_label.setText("Error al generar reporte")
        self._set_controls_enabled(True)

    # ------------------------------------------------------------------
    def _exportar_excel(self) -> None:
        mes, anio = self._get_mes_anio()
        objetivo_id = self._get_objetivo_id()
        nombre_obj = self.selector_objetivo.currentText().replace(" ", "_")
        ruta, _ = QFileDialog.getSaveFileName(
            self, "Guardar Excel",
            f"reporte_{nombre_obj}_{anio}_{mes:02d}.xlsx",
            "Excel (*.xlsx)"
        )
        if ruta:
            self._set_controls_enabled(False)
            self.estado_label.setText("Exportando a Excel...")
            task = run_background_task(_exportar_excel_objetivo, anio, mes, objetivo_id, ruta)
            task.signals.finished.connect(lambda _: self._on_export_exitoso(ruta))
            task.signals.error.connect(self._on_error)

    def _exportar_pdf(self) -> None:
        mes, anio = self._get_mes_anio()
        objetivo_id = self._get_objetivo_id()
        nombre_obj = self.selector_objetivo.currentText().replace(" ", "_")
        ruta, _ = QFileDialog.getSaveFileName(
            self, "Guardar PDF",
            f"reporte_{nombre_obj}_{anio}_{mes:02d}.pdf",
            "PDF (*.pdf)"
        )
        if ruta:
            self._set_controls_enabled(False)
            self.estado_label.setText("Exportando a PDF...")
            task = run_background_task(_exportar_pdf_objetivo, anio, mes, objetivo_id, ruta)
            task.signals.finished.connect(lambda _: self._on_export_exitoso(ruta))
            task.signals.error.connect(self._on_error)

    def _on_export_exitoso(self, ruta: str) -> None:
        QMessageBox.information(self, "Exportación completa", f"Archivo guardado en:\n{ruta}")
        self.estado_label.setText("Exportación completada")
        self._set_controls_enabled(True)