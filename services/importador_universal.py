# =============================================================================
# VESP Organizations - Sistema de Importación Universal
# Soporta Excel, JSON (tablets), y preparado para más formatos
# =============================================================================

import json
import logging
import os
import re
import unicodedata
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from typing import Any, Dict, List, Optional, Tuple, Callable, Set

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger(__name__)

from database.gestor_db import gestor_db
from .gestor_turnos import GestorTurnos
from .sync_manager import get_sync_manager


@dataclass
class RegistroImportacion:
    """Representa un registro a importar."""
    fecha: str
    hora: str
    turno: str
    supervisor: str
    objetivo: str
    fuente: str
    notas: Optional[str] = None
    sheet_title: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            'fecha': self.fecha,
            'hora': self.hora,
            'turno': self.turno,
            'supervisor': self.supervisor,
            'objetivo': self.objetivo,
            'notas': self.notas,
            'fuente': self.fuente,
            'sheet_title': self.sheet_title,
        }


@dataclass
class ResultadoImportacion:
    """Resultado de una importación."""
    total_registros: int
    registros_validos: int
    registros_errores: int
    registros_duplicados: int
    errores: List[str]
    duplicados: List[Dict[str, Any]]
    exitoso: bool


@dataclass
class EvaluacionHojaControl:
    """Resultado estructurado de la evaluación de una hoja del parser de control recorridos."""
    es_valida: bool
    razon: str
    detalle: Optional[str] = None
    fecha: Optional[date] = None
    turno: Optional[str] = None
    tipo: str = 'desconocido'


@dataclass
class WorkbookMetrics:
    total_hojas: int = 0
    hojas_analizadas: int = 0
    hojas_validas: int = 0
    hojas_descartadas: int = 0
    motivos_descarte: Dict[str, int] = None

    def __post_init__(self):
        if self.motivos_descarte is None:
            self.motivos_descarte = {}

    def registrar_descartada(self, motivo: str) -> None:
        self.hojas_descartadas += 1
        self.motivos_descarte[motivo] = self.motivos_descarte.get(motivo, 0) + 1


@dataclass
class RowMetrics:
    total_filas_recorridas: int = 0
    filas_vacias: int = 0
    filas_ocultas: int = 0
    filas_con_error: int = 0
    filas_encabezado: int = 0
    filas_datos: int = 0
    motivos_descarte: Dict[str, int] = None

    def __post_init__(self):
        if self.motivos_descarte is None:
            self.motivos_descarte = {}

    def registrar_descartada(self, motivo: str) -> None:
        self.motivos_descarte[motivo] = self.motivos_descarte.get(motivo, 0) + 1


@dataclass
class ExtractionMetrics:
    objetivos_vacios: int = 0
    supervisores_vacios: int = 0
    horas_vacias: int = 0
    horas_invalidas: int = 0
    turnos_invalidos: int = 0
    registros_creados: int = 0
    registros_descartados: int = 0
    excepciones_parseo: int = 0
    anotaciones_descartadas: int = 0
    motivos_descarte: Dict[str, int] = None

    def __post_init__(self):
        if self.motivos_descarte is None:
            self.motivos_descarte = {}

    def registrar_descartado(self, motivo: str) -> None:
        self.registros_descartados += 1
        self.motivos_descarte[motivo] = self.motivos_descarte.get(motivo, 0) + 1
        self.anotaciones_descartadas += 1

@dataclass
class ValidationMetrics:
    supervisor_inexistente: int = 0
    objetivo_inexistente: int = 0
    fecha_invalida: int = 0
    hora_invalida: int = 0
    turno_invalido: int = 0
    registros_validos: int = 0
    registros_rechazados: int = 0
    motivos_rechazo: Dict[str, int] = None

    def __post_init__(self):
        if self.motivos_rechazo is None:
            self.motivos_rechazo = {}

    def registrar_rechazo(self, motivo: str) -> None:
        self.registros_rechazados += 1
        self.motivos_rechazo[motivo] = self.motivos_rechazo.get(motivo, 0) + 1


@dataclass
class ImportMetrics:
    workbook: WorkbookMetrics = None
    filas: RowMetrics = None
    extraccion: ExtractionMetrics = None
    validacion: ValidationMetrics = None
    importacion: Dict[str, int] = None

    def __post_init__(self):
        if self.workbook is None:
            self.workbook = WorkbookMetrics()
        if self.filas is None:
            self.filas = RowMetrics()
        if self.extraccion is None:
            self.extraccion = ExtractionMetrics()
        if self.validacion is None:
            self.validacion = ValidationMetrics()
        if self.importacion is None:
            self.importacion = {
                'registros_insertados': 0,
                'registros_duplicados': 0,
                'errores_sql': 0,
                'inserciones_fallidas': 0,
            }

    def reset(self) -> None:
        self.workbook = WorkbookMetrics()
        self.filas = RowMetrics()
        self.extraccion = ExtractionMetrics()
        self.validacion = ValidationMetrics()
        self.importacion = {
            'registros_insertados': 0,
            'registros_duplicados': 0,
            'errores_sql': 0,
            'inserciones_fallidas': 0,
        }

    def to_dict(self) -> Dict[str, Any]:
        return {
            'workbook': {
                'total_hojas': self.workbook.total_hojas,
                'hojas_analizadas': self.workbook.hojas_analizadas,
                'hojas_validas': self.workbook.hojas_validas,
                'hojas_descartadas': self.workbook.hojas_descartadas,
                'motivos_descarte': self.workbook.motivos_descarte,
            },
            'filas': {
                'total_filas_recorridas': self.filas.total_filas_recorridas,
                'filas_vacias': self.filas.filas_vacias,
                'filas_ocultas': self.filas.filas_ocultas,
                'filas_con_error': self.filas.filas_con_error,
                'filas_encabezado': self.filas.filas_encabezado,
                'filas_datos': self.filas.filas_datos,
                'motivos_descarte': self.filas.motivos_descarte,
            },
            'extraccion': {
                'objetivos_vacios': self.extraccion.objetivos_vacios,
                'supervisores_vacios': self.extraccion.supervisores_vacios,
                'horas_vacias': self.extraccion.horas_vacias,
                'horas_invalidas': self.extraccion.horas_invalidas,
                'turnos_invalidos': self.extraccion.turnos_invalidos,
                'registros_creados': self.extraccion.registros_creados,
                'registros_descartados': self.extraccion.registros_descartados,
                'excepciones_parseo': self.extraccion.excepciones_parseo,
                'motivos_descarte': self.extraccion.motivos_descarte,
            },
            'validacion': {
                'supervisor_inexistente': self.validacion.supervisor_inexistente,
                'objetivo_inexistente': self.validacion.objetivo_inexistente,
                'fecha_invalida': self.validacion.fecha_invalida,
                'hora_invalida': self.validacion.hora_invalida,
                'turno_invalido': self.validacion.turno_invalido,
                'registros_validos': self.validacion.registros_validos,
                'registros_rechazados': self.validacion.registros_rechazados,
                'motivos_rechazo': self.validacion.motivos_rechazo,
            },
            'importacion': self.importacion,
        }


class ImportadorUniversal:
    """Sistema unificado para importar datos desde múltiples fuentes."""

    _INSTANCIAS: Set["ImportadorUniversal"] = set()
    _ALIAS_ENCABEZADOS: Dict[str, Tuple[str, ...]] = {
        'objetivo': ('objetivo', 'objetivos'),
        'supervisor': ('supervisor', 'supervisores'),
        'hora': ('hora', 'horas'),
        'turno': ('turno', 'turnos'),
        'veces': ('veces', 'cantidad', 'cantidades', 'cant'),
        'notas': ('nota', 'notas', 'observacion', 'observaciones'),
    }
    _PATRONES_ANOTACION = (
        "ingreso a obra",
        "ingresa a obra",
        "se deja garita",
        "garita nro",
        "garita n°",
        "sale movil",
        "sale móvil",
        "llama supervisor",
        "se retira",
        "se entrega",
        "se recibe",
    )

    def __init__(self):
        self.sync_manager = get_sync_manager()

        self._cache_objetivos: Dict[str, int] = {}
        self._cache_supervisores: Dict[str, int] = {}
        self._cache_inicializado = False
        self._metrics = ImportMetrics()
        type(self)._INSTANCIAS.add(self)

    def __del__(self) -> None:
        try:
            type(self)._INSTANCIAS.discard(self)
        except Exception:
            pass

    @classmethod
    def invalidar_cache_global(cls) -> None:
        """Invalida el cache interno de todos los importadores activos."""
        for instancia in list(cls._INSTANCIAS):
            try:
                instancia.invalidate_cache()
            except Exception as exc:
                logger.warning("No se pudo invalidar cache del importador %s: %s", instancia, exc)

    def _reset_cache_state(self) -> None:
        """Resetea el estado interno del cache sin cambiar la API pública."""
        self._cache_objetivos = {}
        self._cache_supervisores = {}
        self._cache_inicializado = False

    def invalidate_cache(self) -> None:
        """Marca el cache como inválido para que se vuelva a cargar en el próximo acceso."""
        self._reset_cache_state()
        logger.info("[CACHE] Cache del importador invalidado")

    def reload_cache(self) -> None:
        """Reconstruye completamente los caches desde la base de datos."""
        self._reset_cache_state()
        self._inicializar_caches()
        logger.info("[CACHE] Cache del importador recargado")

    def _cargar_cache_objetivos(self) -> bool:
        """Carga los objetivos desde la base de datos en el cache interno.

        Returns:
            True si la carga fue exitosa, False si falló (y por lo tanto
            el cache no debe considerarse válido).
        """
        try:
            filas = gestor_db.ejecutar("SELECT id, nombre FROM objetivos")
            for fila in filas:
                key = self._normalizar_texto(fila['nombre'])
                self._cache_objetivos[key] = int(fila['id'])
            logger.info("[CACHE] Cargados %d objetivos", len(self._cache_objetivos))
            return True
        except Exception as e:
            logger.error("No se pudieron cargar objetivos: %s", e)
            if logger.isEnabledFor(logging.DEBUG):
                logger.debug("Cache objetivos traceback", exc_info=True)
            return False
    
    def _cargar_cache_supervisores(self) -> bool:
        """Carga los supervisores desde la base de datos en el cache interno.

        Returns:
            True si la carga fue exitosa, False si falló.
        """
        try:
            filas = gestor_db.ejecutar("SELECT id, nombre FROM supervisores")
            for fila in filas:
                key = self._normalizar_texto(fila['nombre'])
                self._cache_supervisores[key] = int(fila['id'])
            logger.info("[CACHE] Cargados %d supervisores", len(self._cache_supervisores))
            return True
        except Exception as e:
            logger.error("No se pudieron cargar supervisores: %s", e)
            if logger.isEnabledFor(logging.DEBUG):
                logger.debug("Cache supervisores traceback", exc_info=True)
            return False
        
    def _inicializar_caches(self) -> None:
        """Carga los caches si aún no están inicializados o si fueron invalidados.

        IMPORTANTE: si alguna de las dos cargas falla (por ejemplo, error
        de conexión a la base de datos), el cache NO se marca como
        inicializado. Esto evita que el importador quede operando con un
        cache vacío pero "válido", lo que generaría falsos negativos del
        tipo "objetivo no encontrado" hasta el próximo reinicio.
        En ese caso, el próximo llamado a _inicializar_caches() (por
        ejemplo, en el siguiente registro procesado durante la misma
        importación) reintentará la carga automáticamente.
        """
        if self._cache_inicializado:
            return

        ok_objetivos = self._cargar_cache_objetivos()
        ok_supervisores = self._cargar_cache_supervisores()

        self._cache_inicializado = ok_objetivos and ok_supervisores
        
    def _log_exception(
        self,
        method: str,
        sheet_title: Optional[str] = None,
        row_idx: Optional[int] = None,
        extra: Optional[str] = None,
        exc: Optional[Exception] = None,
    ) -> None:
        details = [f"method={method}"]
        if sheet_title:
            details.append(f"sheet={sheet_title}")
        if row_idx is not None:
            details.append(f"row={row_idx}")
        if extra:
            details.append(extra)

        message = " | ".join(details)
        if exc is not None:
            logger.warning("%s: %s", message, exc)
            if logger.isEnabledFor(logging.DEBUG):
                logger.debug("Traceback %s", message, exc_info=True)
        else:
            logger.warning("%s", message)

    def previsualizar_archivo(
        self,
        ruta_archivo: str,
        sheet_names: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """
        Previsualiza un archivo Excel y detecta si es CONTROL_RECORRIDOS.

        NUEVA LÓGICA:
        - Extrae TODOS los objetivos detectados
        - Determina cuáles existen en BD (resueltos) y cuáles no
        - Devuelve mapeos preparados para auto-resolver objetivos existentes
        - El usuario solo ve objetivos NO resueltos que necesitan crear o mapear
        """
        try:
            self._metrics.reset()
            # Inicializar caches una sola vez
            self._inicializar_caches()
            
            # Abrir el workbook con manejo fino de excepciones para dar mensajes
            # accionables al usuario.
            try:
                wb_data = load_workbook(ruta_archivo, data_only=True)
            except FileNotFoundError:
                return {
                    'tipo': 'error',
                    'error': f"Archivo no encontrado: {ruta_archivo}",
                    'registros': [],
                    'objetivos_detectados': [],
                    'supervisores_detectados': [],
                    'objetivos_resueltos': {},
                    'supervisores_resueltos': {},
                    'objetivos_no_resueltos': [],
                    'supervisores_no_resueltos': [],
                    'sheet_options': [],
                }
            except PermissionError:
                return {
                    'tipo': 'error',
                    'error': f"Sin permisos para leer el archivo: {ruta_archivo}",
                    'registros': [],
                    'objetivos_detectados': [],
                    'supervisores_detectados': [],
                    'objetivos_resueltos': {},
                    'supervisores_resueltos': {},
                    'objetivos_no_resueltos': [],
                    'supervisores_no_resueltos': [],
                    'sheet_options': [],
                }
            except InvalidFileException as e:
                return {
                    'tipo': 'error',
                    'error': f"Archivo inválido o corrupto: {e}",
                    'registros': [],
                    'objetivos_detectados': [],
                    'supervisores_detectados': [],
                    'objetivos_resueltos': {},
                    'supervisores_resueltos': {},
                    'objetivos_no_resueltos': [],
                    'supervisores_no_resueltos': [],
                    'sheet_options': [],
                }
            except Exception as e:
                logger.error("Error abriendo Excel: %s", e)
                if logger.isEnabledFor(logging.DEBUG):
                    logger.debug("Traceback abriendo Excel", exc_info=True)
                return {
                    'tipo': 'error',
                    'error': f"Error leyendo Excel: {str(e)}",
                    'registros': [],
                    'objetivos_detectados': [],
                    'supervisores_detectados': [],
                    'objetivos_resueltos': {},
                    'supervisores_resueltos': {},
                    'objetivos_no_resueltos': [],
                    'supervisores_no_resueltos': [],
                    'sheet_options': [],
                }

            # Detectar fórmulas sin valor calculado: comparar con una carga que
            # preserve fórmulas (data_only=False). Si una celda tiene None en la
            # versión data_only pero en la versión con fórmulas la celda contiene
            # una fórmula (empieza con '='), informar al usuario para que abra y
            # guarde el archivo en Excel/LibreOffice.
            try:
                wb_formulas = load_workbook(ruta_archivo, data_only=False)
                problemas = []
                for ws_data, ws_formula in zip(wb_data.worksheets, wb_formulas.worksheets):
                    max_row = min(ws_data.max_row, 40)
                    max_col = min(ws_data.max_column, 20)
                    for r in range(1, max_row + 1):
                        for c in range(1, max_col + 1):
                            val_data = ws_data.cell(row=r, column=c).value
                            val_formula = ws_formula.cell(row=r, column=c).value
                            if (val_data is None or (isinstance(val_data, str) and val_data.strip()=="")) and isinstance(val_formula, str) and val_formula.startswith('='):
                                problemas.append(f"{ws_data.title}:{r}:{c}")
                if problemas:
                    return {
                        'tipo': 'error',
                        'error': (
                            'El archivo contiene celdas con fórmulas sin valor calculado. '
                            'Abrilo en Excel/LibreOffice, guardalo y volvé a subirlo. '
                            f'Ejemplos: {problemas[:5]}'
                        ),
                        'registros': [],
                        'objetivos_detectados': [],
                        'supervisores_detectados': [],
                        'objetivos_resueltos': {},
                        'supervisores_resueltos': {},
                        'objetivos_no_resueltos': [],
                        'supervisores_no_resueltos': [],
                        'sheet_options': [],
                    }
            except Exception:
                # Si falla abrir la versión con fórmulas no abortamos la importación;
                # esto sólo reduce la verificación. Logueamos en debug.
                if logger.isEnabledFor(logging.DEBUG):
                    logger.debug('No se pudo abrir workbook en modo fórmulas para verificar celdas', exc_info=True)

            wb = wb_data

            sheet_options = self._listar_sheet_options(wb)

            # =====================================================
            # Manejo si no hay hojas válidas
            # =====================================================
            if not sheet_options:
                return {
                    'tipo': 'empty',
                    'registros': [],
                    'objetivos_detectados': [],
                    'supervisores_detectados': [],
                    'objetivos_resueltos': {},  # NUEVO: objetivos que SÍ existen
                    'supervisores_resueltos': {},  # NUEVO: supervisores que SÍ existen
                    'objetivos_no_resueltos': [],
                    'supervisores_no_resueltos': [],
                    'sheet_options': [],
                }

            # =====================================================
            # Parsear hojas CONTROL_RECORRIDOS
            # =====================================================
            control = self._parsear_control_recorridos(
                wb,
                sheet_names=sheet_names,
            )

            registros = control.get('registros', [])

            # =====================================================
            # EXTRAER TODOS LOS OBJETIVOS Y SUPERVISORES DETECTADOS
            # =====================================================
            objetivos_detectados = sorted({
                r.objetivo.strip()
                for r in registros
                if r.objetivo and str(r.objetivo).strip()
            })

            supervisores_detectados = sorted({
                r.supervisor.strip()
                for r in registros
                if r.supervisor and str(r.supervisor).strip()
            })

            # =====================================================
            # CLASIFICAR EN RESUELTOS Y NO RESUELTOS
            # =====================================================
            objetivos_resueltos = {}
            objetivos_no_resueltos = []
            
            for nombre in objetivos_detectados:
                objetivo_id = self._obtener_objetivo_id(nombre)
                if objetivo_id is not None:
                    objetivos_resueltos[nombre] = objetivo_id
                else:
                    objetivos_no_resueltos.append(nombre)

            supervisores_resueltos = {}
            supervisores_no_resueltos = []
            
            for nombre in supervisores_detectados:
                supervisor_id = self._obtener_supervisor_id(nombre)
                if supervisor_id is not None:
                    supervisores_resueltos[nombre] = supervisor_id
                else:
                    supervisores_no_resueltos.append(nombre)

            # =====================================================
            # LOG DE RESUMEN
            # =====================================================
            logger.info("[PREVIEW] Archivo: %s", ruta_archivo)
            logger.info("  Total registros: %d", len(registros))
            logger.info(
                "  Objetivos: %d (%d resueltos, %d nuevos)",
                len(objetivos_detectados),
                len(objetivos_resueltos),
                len(objetivos_no_resueltos),
            )
            logger.info(
                "  Supervisores: %d (%d resueltos, %d nuevos)",
                len(supervisores_detectados),
                len(supervisores_resueltos),
                len(supervisores_no_resueltos),
            )

            # =====================================================
            # RESPUESTA FINAL
            # =====================================================
            return {
                'tipo': 'control_recorridos',
                'registros': registros,
                'objetivos_detectados': objetivos_detectados,
                'supervisores_detectados': supervisores_detectados,
                'objetivos_resueltos': objetivos_resueltos,
                'supervisores_resueltos': supervisores_resueltos,
                'objetivos_no_resueltos': objetivos_no_resueltos,
                'supervisores_no_resueltos': supervisores_no_resueltos,
                'sheet_options': sheet_options,
            }

        except Exception as e:
            logger.error("Error en previsualizar_archivo: %s", e)
            if logger.isEnabledFor(logging.DEBUG):
                logger.debug("Traceback en previsualizar_archivo", exc_info=True)
            return {
                'tipo': 'error',
                'error': str(e),
                'registros': [],
                'objetivos_detectados': [],
                'supervisores_detectados': [],
                'objetivos_resueltos': {},
                'supervisores_resueltos': {},
                'objetivos_no_resueltos': [],
                'supervisores_no_resueltos': [],
                'sheet_options': [],
            }
    
    def importar_excel(self, ruta_archivo: str) -> ResultadoImportacion:
        """Importa datos desde archivo Excel."""
        try:
            preview = self.previsualizar_archivo(ruta_archivo)
            if preview.get('tipo') == 'control_recorridos' and len(preview.get('registros', [])) > 0:
                return self.importar_registros(preview['registros'])

            df = pd.read_excel(ruta_archivo, sheet_name='Pasadas')
            columnas_requeridas = ['Fecha', 'Supervisor', 'Objetivo', 'Hora', 'Turno']
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]

            if columnas_faltantes:
                return ResultadoImportacion(
                    total_registros=0,
                    registros_validos=0,
                    registros_errores=0,
                    registros_duplicados=0,
                    errores=[f"Columnas faltantes: {', '.join(columnas_faltantes)}"],
                    duplicados=[],
                    exitoso=False,
                )

            registros = []
            for index, row in df.iterrows():
                try:
                    registro = RegistroImportacion(
                        fecha=self._formatear_fecha(row['Fecha']),
                        hora=self._formatear_hora(row['Hora']),
                        turno=str(row['Turno']).strip().lower(),
                        supervisor=str(row['Supervisor']).strip(),
                        objetivo=str(row['Objetivo']).strip(),
                        notas=str(row.get('Notas', '')) if pd.notna(row.get('Notas')) else None,
                        fuente='excel',
                    )
                    registros.append(registro)
                except Exception as e:
                    logger.warning(
                        "Fila %s de Excel inválida: %s",
                        index,
                        e,
                    )
                    if logger.isEnabledFor(logging.DEBUG):
                        logger.debug(
                            "Traceback fila inválida Excel: %s",
                            e,
                            exc_info=True,
                        )
                    continue

            return self.importar_registros(registros)

        except Exception as e:
            return ResultadoImportacion(
                total_registros=0,
                registros_validos=0,
                registros_errores=0,
                registros_duplicados=0,
                errores=[f"Error leyendo Excel: {str(e)}"],
                duplicados=[],
                exitoso=False,
            )

    def importar_control_recorridos(
        self,
        ruta_archivo: str,
        objetivo_mapeo: Optional[Dict[str, int]] = None,
        supervisor_mapeo: Optional[Dict[str, int]] = None,
        progress_callback: Optional[Callable[[int, int], None]] = None,
        sheet_names: Optional[List[str]] = None,
        sheet_turno_map: Optional[Dict[str, str]] = None,
        preview_precalculado: Optional[Dict[str, Any]] = None,
    ) -> ResultadoImportacion:
        """
        Importa CONTROL_RECORRIDOS con auto-resolución de objetivos y supervisores.
        
        Si objetivo_mapeo o supervisor_mapeo es None, auto-resuelve los que existen en BD.
        Si preview_precalculado se proporciona, lo reutiliza (evita doble parseo).
        """
        
        # Inicializar caches
        self._inicializar_caches()
        
        # Usar preview precalculado si está disponible
        if preview_precalculado is not None:
            preview = preview_precalculado
        else:
            preview = self.previsualizar_archivo(
                ruta_archivo,
                sheet_names=sheet_names,
            )

        if preview.get('tipo') != 'control_recorridos':
            return ResultadoImportacion(
                total_registros=0,
                registros_validos=0,
                registros_errores=1,
                registros_duplicados=0,
                errores=[f"Tipo de archivo inválido: {preview.get('tipo')}"],
                duplicados=[],
                exitoso=False,
            )

        registros = preview.get('registros', [])
        if len(registros) == 0:
            return ResultadoImportacion(
                total_registros=0,
                registros_validos=0,
                registros_errores=1,
                registros_duplicados=0,
                errores=['No se encontraron registros válidos'],
                duplicados=[],
                exitoso=False,
            )

        # ================================================================
        # AUTO-RESOLVER OBJETIVOS Y SUPERVISORES QUE EXISTEN EN BD
        # ================================================================
        
        # Mapeo final que combina auto-resueltos + custom mapeos
        mapeo_objetivo_final = dict(preview.get('objetivos_resueltos', {}))
        if objetivo_mapeo:
            mapeo_objetivo_final.update(objetivo_mapeo)

        mapeo_supervisor_final = dict(preview.get('supervisores_resueltos', {}))
        if supervisor_mapeo:
            mapeo_supervisor_final.update(supervisor_mapeo)

        # ================================================================
        # LOG DE MAPEOS
        # ================================================================
        logger.info("[IMPORTACIÓN] Mapeos finales:")
        logger.info("  Objetivos mapeados: %d", len(mapeo_objetivo_final))
        for nombre, obj_id in sorted(mapeo_objetivo_final.items()):
            logger.debug("    %s -> ID %s", nombre, obj_id)
        logger.info("  Supervisores mapeados: %d", len(mapeo_supervisor_final))
        for nombre, sup_id in sorted(mapeo_supervisor_final.items()):
            logger.debug("    %s -> ID %s", nombre, sup_id)

        # Aplicar mapeo de turnos por sheet si es necesario
        if sheet_turno_map:
            for r in registros:
                mapped = sheet_turno_map.get(r.sheet_title)
                if mapped:
                    r.turno = mapped

        return self.importar_registros(
            registros,
            objetivo_mapeo=mapeo_objetivo_final,
            supervisor_mapeo=mapeo_supervisor_final,
            progress_callback=progress_callback,
        )

    def importar_registros(
        self,
        registros: List[RegistroImportacion],
        objetivo_mapeo: Optional[Dict[str, int]] = None,
        supervisor_mapeo: Optional[Dict[str, int]] = None,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> ResultadoImportacion:
        """Procesa una lista de registros ya normalizados."""

        return self._procesar_registros(
            registros,
            objetivo_mapeo=objetivo_mapeo or {},
            supervisor_mapeo=supervisor_mapeo or {},
            progress_callback=progress_callback,
        )

    def importar_json_tablet(self, ruta_archivo: str, progress_callback: Optional[Callable[[int, int], None]] = None) -> ResultadoImportacion:
        """Importa datos desde archivo JSON de tablet."""
        try:
            with open(ruta_archivo, 'r', encoding='utf-8') as f:
                data = json.load(f)

            registros = []
            for index, pasada in enumerate(data.get('pasadas', []), start=1):
                try:
                    timestamp = datetime.fromisoformat(pasada['timestamp'].replace('Z', '+00:00'))
                    registro = RegistroImportacion(
                        fecha=timestamp.strftime('%Y-%m-%d'),
                        hora=timestamp.strftime('%H:%M'),
                        turno=pasada['turno'],
                        supervisor=data['meta']['usuario'],
                        objetivo=pasada['objetivo'],
                        notas=pasada.get('notas'),
                        fuente='tablet',
                    )
                    registros.append(registro)
                except Exception as e:
                    logger.warning(
                        "Pasada JSON tablet inválida #%s en %s: %s",
                        index,
                        ruta_archivo,
                        e,
                    )
                    if logger.isEnabledFor(logging.DEBUG):
                        logger.debug(
                            "Traceback pasada JSON inválida: %s",
                            e,
                            exc_info=True,
                        )
                    continue

            return self.importar_registros(registros, progress_callback=progress_callback)

        except Exception as e:
            return ResultadoImportacion(
                total_registros=0,
                registros_validos=0,
                registros_errores=0,
                registros_duplicados=0,
                errores=[f"Error leyendo JSON: {str(e)}"],
                duplicados=[],
                exitoso=False,
            )

    def importar_json_string(self, json_string: str) -> ResultadoImportacion:
        """Importa datos desde string JSON (útil para API futuras)."""
        try:
            data = json.loads(json_string)

            tmp_name = None
            try:
                # Crear archivo temporal en el directorio temporal del sistema
                with tempfile.NamedTemporaryFile('w', encoding='utf-8', suffix='.json', delete=False) as tmpf:
                    json.dump(data, tmpf)
                    tmp_name = tmpf.name

                resultado = self.importar_json_tablet(tmp_name)
                return resultado
            finally:
                if tmp_name:
                    try:
                        os.remove(tmp_name)
                    except Exception as e:
                        logger.debug(
                            "No se pudo eliminar archivo temporal %s: %s",
                            tmp_name,
                            e,
                            exc_info=logger.isEnabledFor(logging.DEBUG),
                        )

        except Exception as e:
            return ResultadoImportacion(
                total_registros=0,
                registros_validos=0,
                registros_errores=0,
                registros_duplicados=0,
                errores=[f"Error procesando JSON: {str(e)}"],
                duplicados=[],
                exitoso=False,
            )

    def filtrar_registros_por_rango(
        self,
        registros: List[RegistroImportacion],
        rango_desde: Optional[Tuple[date, str]] = None,
        rango_hasta: Optional[Tuple[date, str]] = None,
    ) -> List[RegistroImportacion]:
        """Filtra registros por un rango de fechas/turnos inclusive."""
        if rango_desde is None and rango_hasta is None:
            return list(registros)

        def _orden_turno(turno: Optional[str]) -> int:
            turno_norm = str(turno or "").strip().lower()
            if turno_norm in ("n", "noche", "nocturno"):
                return 0
            if turno_norm in ("d", "dia", "diurno"):
                return 1
            return 2

        def _comparar_fecha_turno(
            registro_fecha: str,
            registro_turno: str,
            limite_fecha: date | str,
            limite_turno: str,
        ) -> int:
            try:
                reg_date = datetime.strptime(registro_fecha, "%Y-%m-%d").date()
            except Exception as e:
                logger.debug(
                    "Comparación fecha/turno inválida: %s | registro_fecha=%s registro_turno=%s limite_fecha=%s limite_turno=%s",
                    e,
                    registro_fecha,
                    registro_turno,
                    limite_fecha,
                    limite_turno,
                )
                return 0

            # Convertir limite_fecha a date si es string
            if isinstance(limite_fecha, str):
                try:
                    lim_date = datetime.strptime(limite_fecha, "%Y-%m-%d").date()
                except Exception as e:
                    logger.debug(
                        "Comparación fecha/turno inválida para limite_fecha: %s | limite_fecha=%s",
                        e,
                        limite_fecha,
                    )
                    return 0
            else:
                lim_date = limite_fecha

            if reg_date < lim_date:
                return -1
            if reg_date > lim_date:
                return 1

            reg_turno = _orden_turno(registro_turno)
            lim_turno = _orden_turno(limite_turno)

            if reg_turno < lim_turno:
                return -1
            if reg_turno > lim_turno:
                return 1
            return 0

        filtrados: List[RegistroImportacion] = []
        for registro in registros:
            incluir = True

            if rango_desde is not None:
                cmp = _comparar_fecha_turno(
                    registro.fecha,
                    registro.turno,
                    rango_desde[0],
                    rango_desde[1],
                )
                if cmp < 0:
                    incluir = False

            if incluir and rango_hasta is not None:
                cmp = _comparar_fecha_turno(
                    registro.fecha,
                    registro.turno,
                    rango_hasta[0],
                    rango_hasta[1],
                )
                if cmp > 0:
                    incluir = False

            if incluir:
                filtrados.append(registro)

        return filtrados

    def _procesar_registros(
        self,
        registros: List[RegistroImportacion],
        objetivo_mapeo: Optional[Dict[str, int]] = None,
        supervisor_mapeo: Optional[Dict[str, int]] = None,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> ResultadoImportacion:
        """Procesa registros importados con auto-resolución de referencias."""

        total = len(registros)
        validos = 0
        errores = []
        duplicados = []

        objetivo_mapeo = objetivo_mapeo or {}
        supervisor_mapeo = supervisor_mapeo or {}

        # Normalizar las claves de los mapeos proporcionados por el usuario
        # para que coincidan con la normalización usada en el cache interno.
        objetivo_mapeo_norm: Dict[str, int] = {}
        for k, v in objetivo_mapeo.items():
            if k is None:
                continue
            objetivo_mapeo_norm[self._normalizar_texto(k)] = v

        supervisor_mapeo_norm: Dict[str, int] = {}
        for k, v in supervisor_mapeo.items():
            if k is None:
                continue
            supervisor_mapeo_norm[self._normalizar_texto(k)] = v

        procesados = 0
        abort = False

        # Inicializar caches para búsquedas adicionales
        self._inicializar_caches()

        logger.info("[PROCESAMIENTO] Iniciando con %d registros", total)
        logger.info("  Mapeo objetivos: %d entradas", len(objetivo_mapeo))
        logger.info("  Mapeo supervisores: %d entradas", len(supervisor_mapeo))

        for registro in registros:
            try:
                # ============================================================
                # 1. RESOLVER SUPERVISOR
                # ============================================================
                supervisor_id = None
                if registro.supervisor:
                    supervisor_id = supervisor_mapeo_norm.get(self._normalizar_texto(registro.supervisor))

                if supervisor_id is None:
                    # Intentar resolver por caché
                    supervisor_id = self._obtener_supervisor_id(registro.supervisor)

                if supervisor_id is None:
                    self._metrics.validacion.supervisor_inexistente += 1
                    self._metrics.validacion.registrar_rechazo('supervisor_inexistente')
                    errores.append(
                        f"Supervisor no encontrado: '{registro.supervisor}'"
                    )
                    logger.warning(
                        "Registro inválido: supervisor no encontrado: %s | sheet=%s",
                        registro.supervisor,
                        registro.sheet_title,
                    )
                    procesados += 1
                    continue

                # ============================================================
                # 2. RESOLVER OBJETIVO
                # ============================================================
                objetivo_id = None
                if registro.objetivo:
                    objetivo_id = objetivo_mapeo_norm.get(self._normalizar_texto(registro.objetivo))

                if objetivo_id is None:
                    # Intentar resolver por caché
                    objetivo_id = self._obtener_objetivo_id(registro.objetivo)

                if objetivo_id is None:
                    self._metrics.validacion.objetivo_inexistente += 1
                    self._metrics.validacion.registrar_rechazo('objetivo_inexistente')
                    errores.append(
                        f"Objetivo no encontrado: '{registro.objetivo}'"
                    )
                    logger.warning(
                        "Registro inválido: objetivo no encontrado: %s | sheet=%s",
                        registro.objetivo,
                        registro.sheet_title,
                    )
                    procesados += 1
                    continue

                # ============================================================
                # 3. VALIDAR TURNO
                # ============================================================
                turno_normalizado = self._normalizar_turno(registro.turno)
                if turno_normalizado is None:
                    self._metrics.validacion.turno_invalido += 1
                    self._metrics.validacion.registrar_rechazo('turno_invalido')
                    errores.append(
                        f"Turno inválido: '{registro.turno}' "
                        f"(debe ser 'diurno' o 'nocturno')"
                    )
                    logger.warning(
                        "Registro inválido: turno inválido: %s | sheet=%s",
                        registro.turno,
                        registro.sheet_title,
                    )
                    procesados += 1
                    continue

                # ============================================================
                # 4. PARSEAR FECHA Y HORA
                # ============================================================
                try:
                    fecha = datetime.strptime(
                        registro.fecha,
                        "%Y-%m-%d"
                    ).date()

                    if registro.hora:
                        hora = datetime.strptime(
                            registro.hora,
                            "%H:%M"
                        ).time()
                    else:
                        hora = None

                    es_control_recorridos = bool(registro.sheet_title)

                    if es_control_recorridos:
                        fecha_operativa = fecha
                    else:
                        fecha_operativa = (
                            GestorTurnos
                            .calcular_fecha_operativa(
                                fecha,
                                hora,
                                turno_normalizado,
                            )
                        )

                except ValueError as e:
                    self._metrics.validacion.fecha_invalida += 1
                    self._metrics.validacion.hora_invalida += 1
                    self._metrics.validacion.registrar_rechazo('fecha_hora_invalida')
                    errores.append(
                        f"Formato fecha/hora inválido "
                        f"(fecha: {registro.fecha}, hora: {registro.hora}): {e}"
                    )
                    logger.warning(
                        "Registro inválido: fecha/hora no parseable: %s | sheet=%s | row_fecha=%s | row_hora=%s",
                        e,
                        registro.sheet_title,
                        registro.fecha,
                        registro.hora,
                    )
                    procesados += 1
                    continue

                # ============================================================
                # 5. VERIFICAR DUPLICADOS
                # ============================================================
                if self._es_duplicado(
                    supervisor_id,
                    objetivo_id,
                    fecha_operativa.strftime('%Y-%m-%d'),
                    registro.hora,
                    turno_normalizado,
                ):
                    self._metrics.importacion['registros_duplicados'] += 1
                    duplicados.append(registro.to_dict())
                    logger.info(
                        "[DUP] %s %s %s: %s | sheet=%s",
                        fecha_operativa,
                        registro.hora,
                        turno_normalizado,
                        registro.objetivo,
                        registro.sheet_title,
                    )
                    procesados += 1
                    continue

                # ============================================================
                # 6. CREAR PASADA
                # ============================================================
                try:
                    creado = self.sync_manager.crear_pasada_offline(
                        fecha_operativa.strftime('%Y-%m-%d'),
                        registro.hora,
                        turno_normalizado,
                        supervisor_id,
                        objetivo_id,
                        registro.notas,
                        validar_turno=not es_control_recorridos,
                    )

                    if creado:
                        validos += 1
                        self._metrics.validacion.registros_validos += 1
                        self._metrics.importacion['registros_insertados'] += 1
                        logger.info(
                            "Pasada creada: %s %s %s | objetivo=%s | supervisor=%s",
                            fecha_operativa,
                            registro.hora,
                            turno_normalizado,
                            registro.objetivo,
                            registro.supervisor,
                        )
                    else:
                        self._metrics.importacion['inserciones_fallidas'] += 1
                        self._metrics.validacion.registrar_rechazo('insercion_fallida')
                        errores.append(
                            f"No se pudo crear pasada: "
                            f"{registro.supervisor} - {registro.objetivo}"
                        )
                        logger.warning(
                            "No se pudo crear pasada: %s | sheet=%s",
                            registro.objetivo,
                            registro.sheet_title,
                        )

                except Exception as e:
                    self._metrics.importacion['errores_sql'] += 1
                    self._metrics.validacion.registrar_rechazo('error_sql')
                    errores.append(
                        f"Error creando pasada: {str(e)}"
                    )
                    logger.error(
                        "Exception al crear pasada: %s | sheet=%s | row_fecha=%s | row_hora=%s",
                        e,
                        registro.sheet_title,
                        registro.fecha,
                        registro.hora,
                    )
                    if logger.isEnabledFor(logging.DEBUG):
                        logger.debug("Traceback crear pasada", exc_info=True)

            except Exception as e:
                self._metrics.extraccion.excepciones_parseo += 1
                self._metrics.extraccion.registrar_descartado('excepcion_parseo')
                errores.append(
                    f"Error procesando registro: {str(e)}"
                )
                logger.error(
                    "Exception procesando registro: %s | sheet=%s | registro=%s",
                    e,
                    registro.sheet_title,
                    registro.to_dict(),
                )
                if logger.isEnabledFor(logging.DEBUG):
                    logger.debug("Traceback registro", exc_info=True)

            finally:
                procesados += 1

                try:
                    if progress_callback:
                        progress_callback(procesados, total)
                except Exception as e:
                    errores.append(
                        f"Importación cancelada: {e}"
                    )
                    logger.error(
                        "Error en progress_callback: %s | processed=%d total=%d",
                        e,
                        procesados,
                        total,
                    )
                    if logger.isEnabledFor(logging.DEBUG):
                        logger.debug("Traceback progress_callback", exc_info=True)
                    abort = True

            if abort:
                break

        # ================================================================
        # RESUMEN FINAL
        # ================================================================
        logger.info("[RESUMEN FINAL]")
        logger.info("[METRICS] %s", self._metrics.to_dict())
        logger.info("  Total: %d", total)
        logger.info("  Válidos: %d", validos)
        logger.info("  Errores: %d", len(errores))
        logger.info("  Duplicados: %d", len(duplicados))

        return ResultadoImportacion(
            total_registros=total,
            registros_validos=validos,
            registros_errores=len(errores),
            registros_duplicados=len(duplicados),
            errores=errores,
            duplicados=duplicados,
            exitoso=(len(errores) == 0 and validos > 0),
        )

    def _parsear_control_recorridos(
        self,
        workbook_or_path,
        sheet_names: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """Parsea un workbook CONTROL_RECORRIDOS y devuelve registros normalizados."""

        if isinstance(workbook_or_path, str):
            workbook = load_workbook(workbook_or_path, data_only=True)
        else:
            workbook = workbook_or_path

        registros = []
        sheet_names_set = set(sheet_names) if sheet_names else None
        self._metrics.workbook.total_hojas = len(workbook.worksheets)

        for ws in workbook.worksheets:
            self._metrics.workbook.hojas_analizadas += 1
            try:
                evaluacion = self._evaluar_hoja_control_recorridos(ws, sheet_names_set)
                if not evaluacion.es_valida:
                    self._metrics.workbook.registrar_descartada(evaluacion.razon)
                    logger.info(
                        "Hoja descartada durante parseo: %s | razon=%s | detalle=%s",
                        ws.title,
                        evaluacion.razon,
                        evaluacion.detalle,
                    )
                    continue

                self._metrics.workbook.hojas_validas += 1
                fecha_base = evaluacion.fecha or date.today()
                turno_base = evaluacion.turno

                hoja_registros = self._parsear_hoja_control_recorridos(
                    ws,
                    fecha_base,
                    turno_base,
                )

                if not hoja_registros:
                    self._metrics.extraccion.registrar_descartado('sin_registros')
                    logger.info(
                        "Hoja sin registros válidos: %s | razon=%s | detalle=%s",
                        ws.title,
                        evaluacion.razon,
                        evaluacion.detalle,
                    )
                    continue

                registros.extend(hoja_registros)
            except Exception as e:
                self._metrics.workbook.registrar_descartada('error_analisis_hoja')
                self._metrics.extraccion.excepciones_parseo += 1
                logger.warning(
                    "Error durante el análisis de hoja: sheet=%s | error=%s",
                    ws.title,
                    e,
                )
                if logger.isEnabledFor(logging.DEBUG):
                    logger.debug(
                        "Traceback hoja inválida: %s",
                        ws.title,
                        exc_info=True,
                    )
                continue

        return {
            'tipo': 'control_recorridos',
            'registros': registros,
            'objetivos_no_resueltos': [],
            'supervisores_no_resueltos': [],
        }

    def _parsear_hoja_control_recorridos(
        self,
        ws,
        sheet_date: date,
        turno: str
    ) -> List[RegistroImportacion]:
        """Parser principal de CONTROL_RECORRIDOS.

        Soporta el formato legacy de 3 bloques horizontales y, si la hoja
        tiene encabezados tipo Objetivo/Supervisor/Hora, también el formato
        tabular simple.
        """

        header_row, columnas = self._buscar_encabezados_control(ws)
        if header_row is not None and columnas:
            return self._parsear_con_encabezados_control(
                ws,
                sheet_date,
                turno,
                columnas,
                header_row,
            )

        return self._parsear_control_recorridos_legacy(
            ws,
            sheet_date,
            turno,
        )

    def _buscar_encabezados_control(self, ws) -> Tuple[Optional[int], Dict[str, Optional[int]]]:
        """Busca la fila de encabezado más probable en la hoja."""
        max_col = ws.max_column
        max_row = min(ws.max_row, 40)
        mejor_fila = None
        mejor_columnas = {}
        mejor_score = 0

        for fila in range(1, max_row + 1):
            columnas = {
                'objetivo': None,
                'supervisor': None,
                'hora': None,
                'turno': None,
                'veces': None,
                'notas': None,
            }

            for columna in range(1, max_col + 1):
                valor = self._limpiar_valor(ws.cell(row=fila, column=columna).value)
                if not valor:
                    continue

                tipo_encabezado = self._identificar_tipo_encabezado(valor)
                if tipo_encabezado == 'objetivo' and columnas['objetivo'] is None:
                    columnas['objetivo'] = columna
                elif tipo_encabezado == 'supervisor' and columnas['supervisor'] is None:
                    columnas['supervisor'] = columna
                elif tipo_encabezado == 'hora' and columnas['hora'] is None:
                    columnas['hora'] = columna
                elif tipo_encabezado == 'turno' and columnas['turno'] is None:
                    columnas['turno'] = columna
                elif tipo_encabezado == 'veces' and columnas['veces'] is None:
                    columnas['veces'] = columna
                elif tipo_encabezado == 'notas' and columnas['notas'] is None:
                    columnas['notas'] = columna

            score = sum(1 for valor in columnas.values() if valor is not None)
            if score > mejor_score and columnas['objetivo'] is not None and columnas['hora'] is not None:
                mejor_score = score
                mejor_fila = fila
                mejor_columnas = columnas

        if mejor_fila is None or mejor_columnas['objetivo'] is None or mejor_columnas['hora'] is None:
            return None, {}

        return mejor_fila, mejor_columnas

    def _resolver_turno_excel(self, turno_raw: Any, turno_fallback: Optional[str]) -> Optional[str]:
        """Resuelve el turno desde la celda de Excel o usa el turno de la hoja como fallback."""
        if turno_raw is None:
            return turno_fallback

        texto = self._limpiar_valor(turno_raw)
        if not texto:
            return turno_fallback

        # Devolver la normalización concreta si se reconoce, sino devolver
        # el fallback (que puede ser None para indicar "no reconocido").
        turno_norm = self._normalizar_turno(texto)
        if turno_norm is not None:
            return turno_norm

        return turno_fallback

    def _normalizar_turno(self, valor: Any) -> Optional[str]:
        """Normaliza variantes textuales de turno a 'diurno' o 'nocturno'.

        Devuelve `None` si no reconoce el valor.
        """
        if valor is None:
            return None

        texto = str(valor).strip().lower()
        if not texto:
            return None

        diurno_aliases = {
            'd', 'dia', 'diurno', 'diaria', 'diario', 'day', 'dayshift', 'matutino'
        }
        nocturno_aliases = {
            'n', 'noche', 'nocturno', 'night', 'nightshift'
        }

        if texto in diurno_aliases:
            return 'diurno'
        if texto in nocturno_aliases:
            return 'nocturno'

        return None


    def _es_anotacion_supervisor(self, texto: str) -> bool:
        """Determina si el texto corresponde a una anotación del supervisor
        y no a un objetivo válido.
        """
        if not texto:
            return False

        texto = self._normalizar_texto(texto)

        # Para evitar falsos positivos, sólo considerar anotaciones cuando
        # el texto es exactamente la frase esperada o cuando la frase aparece
        # al inicio y el resto del contenido es corto (p.ej. una hora).
        for patron in self._PATRONES_ANOTACION:
            patron_norm = self._normalizar_texto(patron)
            if texto == patron_norm:
                return True

            # Caso común: "SE RETIRA A LAS 22:00" -> comienza con la frase
            if texto.startswith(patron_norm + ' '):
                resto = texto[len(patron_norm):].strip()
                # Si el resto es vacío, es una anotación.
                if not resto:
                    return True
                # Si el resto contiene dígitos (horas) y es corto, considerarlo anotación.
                if re.search(r"\d", resto) and len(resto) <= 25:
                    return True

        return False

    def _crear_registros_control_recorridos(
        self,
        objetivo_raw: Any,
        supervisor_raw: Any,
        hora_raw: Any,
        turno_raw: Any,
        notas_raw: Any,
        veces_raw: Any,
        sheet_date: date,
        turno_hoja: str,
        ws_title: str,
        row_idx: Optional[int] = None,
        bloque_idx: Optional[int] = None,
        origen: str = 'parser',
    ) -> List[RegistroImportacion]:
        """Normaliza y construye registros de CONTROL_RECORRIDOS desde datos crudos."""
        objetivo = self._limpiar_valor(objetivo_raw)
        if not objetivo:
            self._metrics.extraccion.objetivos_vacios += 1
            self._metrics.extraccion.registrar_descartado('objetivo_vacio')
            return []

        if self._es_anotacion_supervisor(objetivo):
            logger.info(
                "Anotación descartada: %s | sheet=%s | row=%s",
                objetivo,
                ws_title,
                row_idx,
            )
            self._metrics.extraccion.registrar_descartado("anotacion_supervisor")
            return []

        if self._es_encabezado(objetivo):
            self._metrics.filas.filas_encabezado += 1
            self._metrics.filas.registrar_descartada('encabezado')
            self._metrics.extraccion.registrar_descartado('encabezado')
            return []
        supervisor = self._limpiar_valor(supervisor_raw) or ''
        if not supervisor:
            self._metrics.extraccion.supervisores_vacios += 1
            self._metrics.extraccion.registrar_descartado('supervisor_vacio')

        notas = self._limpiar_valor(notas_raw) if notas_raw is not None else None
        turno_registro = self._resolver_turno_excel(turno_raw, turno_hoja)
        if turno_raw is not None:
            texto_turno = self._limpiar_valor(turno_raw)
            if texto_turno and turno_registro is None:
                self._metrics.extraccion.turnos_invalidos += 1
                self._metrics.extraccion.registrar_descartado('turno_invalido')
            elif texto_turno and turno_registro == turno_hoja and not self._resolver_turno_excel(texto_turno, None):
                self._metrics.extraccion.turnos_invalidos += 1
                self._metrics.extraccion.registrar_descartado('turno_invalido')
        if turno_registro is None:
            turno_registro = turno_hoja

        repeticiones = self._parsear_repeticiones(veces_raw)
        if repeticiones <= 0:
            repeticiones = 1

        fecha_import = sheet_date
        hora_normalizada = ''

        if hora_raw is None or str(hora_raw).strip() == '':
            self._metrics.extraccion.horas_vacias += 1
            self._metrics.extraccion.registrar_descartado('hora_vacia')
        else:
            try:
                fecha_import, hora_normalizada = self._normalizar_hora_y_fecha(hora_raw, sheet_date)
            except Exception as e:
                self._metrics.extraccion.horas_invalidas += 1
                self._metrics.extraccion.registrar_descartado('hora_invalida')
                logger.warning(
                    "Hora inválida en control recorridos (%s): %s | sheet=%s | row=%s | bloque=%s | objetivo=%s | hora=%s",
                    origen,
                    e,
                    ws_title,
                    row_idx,
                    bloque_idx,
                    objetivo,
                    hora_raw,
                )
                if logger.isEnabledFor(logging.DEBUG):
                    logger.debug(
                        "Traceback hora inválida control recorridos: %s",
                        ws_title,
                        exc_info=True,
                    )
                hora_normalizada = ''

        registros = []
        for _ in range(repeticiones):
            registros.append(
                RegistroImportacion(
                    fecha=fecha_import.strftime('%Y-%m-%d'),
                    hora=hora_normalizada,
                    turno=turno_registro or turno_hoja,
                    supervisor=supervisor,
                    objetivo=objetivo,
                    notas=notas,
                    fuente='excel',
                    sheet_title=ws_title,
                )
            )

        self._metrics.extraccion.registros_creados += len(registros)
        return registros

    def _parsear_con_encabezados_control(
        self,
        ws,
        sheet_date: date,
        turno: str,
        columnas: Dict[str, Optional[int]],
        header_row: int,
    ) -> List[RegistroImportacion]:
        registros = []
        max_row = ws.max_row
        max_col = ws.max_column

        self._metrics.filas.total_filas_recorridas += 1
        if header_row is not None:
            self._metrics.filas.filas_encabezado += 1
            self._metrics.filas.registrar_descartada('encabezado')

        for fila in range(header_row + 1, max_row + 1):
            self._metrics.filas.total_filas_recorridas += 1
            if ws.row_dimensions[fila].hidden:
                self._metrics.filas.filas_ocultas += 1
                self._metrics.filas.registrar_descartada('fila_oculta')
                continue

            fila_valores = [
                ws.cell(row=fila, column=columna).value
                for columna in range(1, max_col + 1)
            ]
            if all(self._limpiar_valor(valor) is None for valor in fila_valores):
                self._metrics.filas.filas_vacias += 1
                self._metrics.filas.registrar_descartada('fila_vacia')
                continue

            valores_relevantes = []
            for clave in ('objetivo', 'supervisor', 'hora', 'turno', 'veces', 'notas'):
                if columnas.get(clave) is not None:
                    valores_relevantes.append(ws.cell(row=fila, column=columnas[clave]).value)

            objetivo_valor = self._limpiar_valor(valores_relevantes[0]) if valores_relevantes else None
            otros_valores = [self._limpiar_valor(valor) for valor in valores_relevantes[1:]]

            if not any(valor is not None and valor != '' for valor in otros_valores) and (objetivo_valor is None or objetivo_valor == ''):
                self._metrics.filas.filas_vacias += 1
                self._metrics.filas.registrar_descartada('fila_vacia')
                continue

            objetivo_raw = ws.cell(row=fila, column=columnas['objetivo']).value
            if objetivo_raw is None:
                self._metrics.filas.registrar_descartada('objetivo_nulo')
                continue

            objetivo_limpio = self._limpiar_valor(objetivo_raw)
            if not objetivo_limpio:
                self._metrics.filas.filas_vacias += 1
                self._metrics.filas.registrar_descartada('fila_vacia')
                self._metrics.extraccion.objetivos_vacios += 1
                self._metrics.extraccion.registrar_descartado('objetivo_vacio')
                continue

            if self._es_encabezado(objetivo_limpio):
                self._metrics.filas.filas_encabezado += 1
                self._metrics.filas.registrar_descartada('encabezado')
                continue

            if fila == header_row:
                self._metrics.filas.filas_encabezado += 1
                self._metrics.filas.registrar_descartada('encabezado')
                continue

            otros_valores = [
                self._limpiar_valor(ws.cell(row=fila, column=columnas[clave]).value)
                if columnas.get(clave) is not None else None
                for clave in ('supervisor', 'hora', 'turno', 'veces', 'notas')
            ]
            if all(valor is None or valor == '' for valor in otros_valores):
                self._metrics.filas.filas_vacias += 1
                self._metrics.filas.registrar_descartada('fila_vacia')
            else:
                self._metrics.filas.filas_datos += 1

            registros.extend(
                self._crear_registros_control_recorridos(
                    objetivo_raw=objetivo_raw,
                    supervisor_raw=(
                        ws.cell(row=fila, column=columnas['supervisor']).value
                        if columnas['supervisor'] is not None else ''
                    ),
                    hora_raw=ws.cell(row=fila, column=columnas['hora']).value,
                    turno_raw=(
                        ws.cell(row=fila, column=columnas['turno']).value
                        if columnas.get('turno') is not None else None
                    ),
                    notas_raw=(
                        ws.cell(row=fila, column=columnas['notas']).value
                        if columnas['notas'] is not None else None
                    ),
                    veces_raw=(
                        ws.cell(row=fila, column=columnas['veces']).value
                        if columnas['veces'] is not None else None
                    ),
                    sheet_date=sheet_date,
                    turno_hoja=turno,
                    ws_title=ws.title,
                    row_idx=fila,
                    origen='encabezados',
                )
            )

        return registros

    def _es_fila_encabezado_global_control_recorridos(self, fila) -> bool:
        """Detecta filas iniciales que contienen títulos generales o cabeceras del formato CONTROL_RECORRIDOS."""
        for valor in fila:
            texto = self._limpiar_valor(valor)
            if not texto:
                continue

            normalizado = self._normalizar_texto(texto)
            if any(keyword in normalizado for keyword in ('control', 'fecha', 'objetivo', 'turno', 'movil', 'supervisor')):
                return True

        return False
    
    def _parsear_control_recorridos_legacy(
        self,
        ws,
        sheet_date: date,
        turno: str
    ) -> List[RegistroImportacion]:
        """
        Parsea el formato CONTROL_RECORRIDOS real con 3 bloques horizontales.

        Cada bloque representa una pasada adicional al mismo objetivo.

        Estructura por bloque:
        NO | OBJETIVO | TURNO | MOVIL | HORA | SUPERVISOR

        Bloques:
        - cols 1-6   -> primera pasada
        - cols 8-13  -> segunda pasada
        - cols 15-20 -> tercera pasada

        Reglas:
        - El turno REAL válido es el de la hoja.
        - Si la fila tiene turno explícito y no coincide con la hoja:
        se ignora.
        - Si la fila no tiene turno:
        se usa el turno de la hoja como fallback.
        """

        # 're' ya se importa a nivel de módulo; import local eliminado.

        registros = []
        filas_procesadas = 0
        filas_salidas = 0
        filas_con_error = 0

        for row_idx, fila in enumerate(
            ws.iter_rows(
                min_row=1,
                max_row=ws.max_row,
                values_only=True,
            ),
            start=1,
        ):

            # ======================================================
            # FIX: Ignorar filas ocultas por filtros de Excel
            # ======================================================

            self._metrics.filas.total_filas_recorridas += 1

            if ws.row_dimensions[row_idx].hidden:
                self._metrics.filas.filas_ocultas += 1
                self._metrics.filas.registrar_descartada('fila_oculta')
                filas_salidas += 1
                continue

            # ======================================================
            # FIX: Ignorar filas completamente vacías
            # ======================================================

            if not fila:
                self._metrics.filas.filas_vacias += 1
                self._metrics.filas.registrar_descartada('fila_vacia')
                filas_salidas += 1
                continue

            if all(
                self._limpiar_valor(valor) is None
                for valor in fila
            ):
                self._metrics.filas.filas_vacias += 1
                self._metrics.filas.registrar_descartada('fila_vacia')
                filas_salidas += 1
                continue

            # Procesar cada bloque en la fila
            for bloque_idx, (c_obj, c_turno, c_hora, c_sup) in enumerate(self._bloques_control_recorridos()):

                idx_obj = c_obj - 1
                idx_turno = c_turno - 1
                idx_hora = c_hora - 1
                idx_sup = c_sup - 1

                # Saltar si la fila no tiene ni siquiera la columna de objetivo
                if len(fila) <= idx_obj:
                    continue

                try:
                    objetivo = fila[idx_obj] if len(fila) > idx_obj else None
                    turno_raw = (
                        fila[idx_turno]
                        if len(fila) > idx_turno
                        else None
                    )
                    hora_raw = fila[idx_hora] if len(fila) > idx_hora else None
                    supervisor = (
                        fila[idx_sup]
                        if len(fila) > idx_sup
                        else None
                    )

                    nuevos_registros = self._crear_registros_control_recorridos(
                        objetivo_raw=objetivo,
                        supervisor_raw=supervisor,
                        hora_raw=hora_raw,
                        turno_raw=turno_raw,
                        notas_raw=None,
                        veces_raw=None,
                        sheet_date=sheet_date,
                        turno_hoja=turno,
                        ws_title=ws.title,
                        row_idx=row_idx,
                        bloque_idx=bloque_idx + 1,
                        origen='legacy',
                    )
                    registros.extend(nuevos_registros)
                    if nuevos_registros:
                        filas_procesadas += 1
                        self._metrics.filas.filas_datos += 1
                    else:
                        self._metrics.extraccion.registrar_descartado('sin_registros_generados')

                except Exception as e:
                    logger.warning(
                        "Error en bloque de control recorridos: %s | sheet=%s | fila=%d | bloque=%d",
                        e,
                        ws.title,
                        row_idx,
                        bloque_idx + 1,
                    )
                    if logger.isEnabledFor(logging.DEBUG):
                        logger.debug(
                            "Traceback bloque control recorridos: %s | sheet=%s | fila=%d | bloque=%d",
                            e,
                            ws.title,
                            row_idx,
                            bloque_idx + 1,
                            exc_info=True,
                        )
                    self._metrics.filas.filas_con_error += 1
                    self._metrics.filas.registrar_descartada('error_parseo_fila')
                    filas_con_error += 1

        logger.info(
            "[PARSE SUMMARY] Hoja: %s | Registros: %d | Filas procesadas: %d | Filas salidas: %d | Filas con error: %d",
            ws.title,
            len(registros),
            filas_procesadas,
            filas_salidas,
            filas_con_error,
        )

        return registros

    def _listar_sheet_options(self, workbook_or_path):
        logger.debug('\n====================')
        logger.debug('DEBUG LISTAR SHEETS')
        logger.debug('====================')

        if isinstance(workbook_or_path, str):
            logger.debug('Abriendo workbook: %s', workbook_or_path)
            workbook = load_workbook(workbook_or_path, data_only=True)
        else:
            workbook = workbook_or_path

        logger.debug('Total hojas: %d', len(workbook.worksheets))

        opciones = []

        for ws in workbook.worksheets:
            logger.debug('HOJA RAW: %r', ws.title)

            try:
                evaluacion = self._evaluar_hoja_control_recorridos(ws)
                if not evaluacion.es_valida:
                    logger.debug('IGNORADA hoja %s | razon=%s | detalle=%s', ws.title, evaluacion.razon, evaluacion.detalle)
                    continue

                opciones.append(
                    {
                        'title': ws.title,
                        'fecha': (evaluacion.fecha or date.today()).isoformat(),
                        'turno': evaluacion.turno,
                    }
                )

                logger.debug('ACEPTADA hoja %s | razon=%s', ws.title, evaluacion.razon)
            except Exception as e:
                logger.warning(
                    "Error en hoja al listar opciones: %s | sheet=%s",
                    e,
                    ws.title,
                )
                if logger.isEnabledFor(logging.DEBUG):
                    logger.debug("Traceback listar opciones hoja %s", ws.title, exc_info=True)

        logger.debug('RESULTADO FINAL: %s', opciones)

        return opciones

    def _bloques_control_recorridos(self) -> List[Tuple[int, int, int, int]]:
        """
        Define los bloques de columnas del formato CONTROL_RECORRIDOS real.
        Cada bloque tiene: NO | OBJETIVO | TURNO | MOVIL | HORA | SUPERVISOR
        Retorna tuplas (col_objetivo, col_turno, col_hora, col_supervisor) — 1-indexed.
        Bloque 1: cols 1-6, Bloque 2: cols 8-13, Bloque 3: cols 15-20.
        """
        return [
            (2, 3, 5, 6),    # bloque 1: cols 1-6
            (9, 10, 12, 13), # bloque 2: cols 8-13
            (16, 17, 19, 20), # bloque 3: cols 15-20
        ]

    def _analizar_nombre_hoja(self, sheet_name: str) -> Dict[str, Any]:
        """Analiza únicamente el nombre de la hoja para extraer metadatos de fecha y turno."""
        texto = str(sheet_name).strip().upper()
        if not texto:
            return {'es_valido': False, 'razon': 'nombre inválido', 'fecha': None, 'turno': None}

        match = re.search(r'(\d{1,2})\s*[-/]\s*(\d{1,2})', texto)
        if not match:
            return {'es_valido': False, 'razon': 'nombre inválido', 'fecha': None, 'turno': None}

        dia = int(match.group(1))
        mes = int(match.group(2))

        turno = None
        if re.search(r'\(D\)|DIURNO|\bD\b', texto):
            turno = 'diurno'
        elif re.search(r'\(N\)|NOCTURNO|\bN\b', texto):
            turno = 'nocturno'

        if turno is None:
            return {'es_valido': False, 'razon': 'turno inexistente', 'fecha': None, 'turno': None}

        YEAR_IMPORTACION = date.today().year
        try:
            fecha = date(YEAR_IMPORTACION, mes, dia)
        except ValueError:
            return {'es_valido': False, 'razon': 'fecha inválida', 'fecha': None, 'turno': turno}

        return {'es_valido': True, 'razon': 'nombre válido', 'fecha': fecha, 'turno': turno}

    def _evaluar_hoja_control_recorridos(
        self,
        ws,
        sheet_names_set: Optional[Set[str]] = None,
    ) -> EvaluacionHojaControl:
        """Evalúa una hoja usando un único criterio para decidir si es importable."""
        try:
            title = self._limpiar_valor(ws.title)
            if not title:
                return EvaluacionHojaControl(False, 'nombre inválido', 'la hoja no tiene nombre')

            if sheet_names_set is not None and title not in sheet_names_set:
                return EvaluacionHojaControl(False, 'sheet no seleccionado', 'la hoja no fue solicitada para procesar')

            if self._hoja_esta_vacia(ws):
                return EvaluacionHojaControl(False, 'hoja vacía', 'la hoja no contiene datos')

            nombre_analizado = self._analizar_nombre_hoja(title)
            if nombre_analizado['es_valido']:
                return EvaluacionHojaControl(
                    True,
                    'nombre válido',
                    'el nombre de la hoja contiene fecha y turno válidos',
                    fecha=nombre_analizado['fecha'],
                    turno=nombre_analizado['turno'],
                    tipo='control_recorridos',
                )

            # Detectar nombres legend que empiezan con 'control' (p.ej. "Control Recorridos").
            # Evitar coincidencias en cualquier parte del título (p.ej. "Panel de Control").
            if re.search(r'^\s*control(?:\b|[_\-\s:\/])', title, re.IGNORECASE):
                return EvaluacionHojaControl(
                    True,
                    'formato control recorridos',
                    'el nombre indica un formato legacy de control recorridos',
                    fecha=date.today(),
                    turno=None,
                    tipo='control_recorridos',
                )

            header_row, columnas = self._buscar_encabezados_control(ws)
            if header_row is not None and columnas:
                return EvaluacionHojaControl(
                    True,
                    'estructura compatible',
                    'la hoja contiene encabezados de control reconocibles',
                    fecha=date.today(),
                    turno=None,
                    tipo='control_recorridos',
                )

            return EvaluacionHojaControl(
                False,
                'formato no reconocido',
                'no se encontraron metadatos de nombre ni estructura compatible',
            )
        except Exception as exc:
            return EvaluacionHojaControl(
                False,
                'error durante el análisis',
                str(exc),
            )

    def _hoja_esta_vacia(self, ws) -> bool:
        """Determina si una hoja está vacía o sin datos útiles."""
        try:
            if ws.max_row <= 1 and ws.max_column <= 1:
                return True

            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 5), values_only=True):
                if any(self._limpiar_valor(valor) is not None for valor in row):
                    return False

            return True
        except Exception:
            return True

    def _parsear_nombre_sheet(self, sheet_name: str):
        """
        Extrae fecha y turno desde nombres tipo:

        01-05 D
        01/05 N
        01-05 (D)
        01-05 DIURNO
        01-05 NOCTURNO

        IMPORTANTE:
        Todas las hojas CONTROL_RECORRIDOS pertenecen
        al año operativo actual configurado por la aplicación.

        No intenta adivinar años.
        """
        # Este método se mantiene como wrapper por compatibilidad con pruebas
        # y código legado que históricamente llamaba a `_parsear_nombre_sheet`.
        # La lógica real está en `_analizar_nombre_hoja` y aquí delegamos en ella.
        analisis = self._analizar_nombre_hoja(sheet_name)
        return analisis.get('fecha'), analisis.get('turno')

    def _normalizar_hora_y_fecha(
        self,
        hora_raw: Any,
        fecha_base: date
    ) -> Tuple[date, str]:
        """
        Normaliza horas provenientes de Excel / Google Sheets / inputs sucios.

        Soporta:
        - 2149 -> 21:49
        - 205 -> 02:05
        - 14 -> 00:14
        - 5 -> 00:05
        - 2149.0 -> 21:49
        - 21:49 / 21;49 / 21.49
        - 03: / 03; -> 03:00
        - :30 / ;30 -> 00:30
        - 26:30 -> día siguiente 02:30
        - datetime.time
        - datetime.datetime
        """

        # 'timedelta' y 'time' ya se importan a nivel de módulo; imports locales eliminados.

        # =========================
        # 1. TIME directo
        # =========================
        if isinstance(hora_raw, time):
            return fecha_base, hora_raw.strftime("%H:%M")
        
        # =========================
        # 2. DATETIME directo
        # =========================
        if isinstance(hora_raw, datetime):
            return fecha_base, hora_raw.strftime("%H:%M")

        # =========================
        # 3. Normalizar a string base
        # =========================
        if hora_raw is None:
            raise ValueError("Hora vacía")

        if isinstance(hora_raw, float):
            if hora_raw.is_integer():
                texto = str(int(hora_raw))
            else:
                texto = str(hora_raw)
        else:
            texto = str(hora_raw)

        # =========================
        # 4. Limpieza fuerte
        # =========================
        texto = texto.strip().lower()

        texto = (
            texto.replace("\u00a0", " ")
                .replace("h", "")
                .replace("hora", "")
                .replace(" ", "")
        )

        # Unificar separadores raros y secuencias de separadores
        texto = texto.replace(";", ":").replace(".", ":")
        texto = re.sub(r":{2,}", ":", texto)

        # =========================
        # 5. Casos vacíos
        # =========================
        if texto in ("", "none", "n/a", "na", "null", "-", "--", "cerrado", "closed"):
            raise ValueError("Hora vacía o inválida")

        # =========================
        # 6. Normalizaciones especiales
        # =========================

        # ":30" o "30:" o ";30"
        if re.fullmatch(r":\d{1,2}", texto):
            texto = "0" + texto

        if re.fullmatch(r"\d{1,2}:", texto):
            texto = texto + "00"

        # =========================
        # 7. Parsing
        # =========================
        hora = None
        minuto = None

        # SOLO MINUTOS: "12"
        if re.fullmatch(r"\d{1,2}", texto):
            hora = 0
            minuto = int(texto)

        # HHMM: 2149 / 205
        elif re.fullmatch(r"\d{3,4}", texto):
            if len(texto) == 3:
                hora = int(texto[0])
                minuto = int(texto[1:])
            else:
                hora = int(texto[:2])
                minuto = int(texto[2:])

        # HH:MM
        elif re.fullmatch(r"\d{1,2}:\d{1,2}", texto):
            hora, minuto = map(int, texto.split(":"))

        elif re.fullmatch(r"\d{1,2}:\d{1,2}:\d{1,2}", texto):
            hora, minuto, _ = map(int, texto.split(":"))

        else:
            raise ValueError(f"Formato de hora no reconocido: '{hora_raw}' (procesado como '{texto}')")

        # =========================
        # 8. Validaciones
        # =========================
        if minuto < 0 or minuto >= 60:
            raise ValueError(f"Minutos inválidos: {minuto} (entrada: {hora_raw})")

        if hora < 0:
            raise ValueError(f"Hora inválida: {hora} (entrada: {hora_raw})")

        # =========================
        # 9. Overflow (26:30 -> día siguiente 02:30)
        # =========================
        total_minutos = hora * 60 + minuto

        dias = total_minutos // (24 * 60)
        total_minutos = total_minutos % (24 * 60)

        fecha_final = fecha_base + timedelta(days=dias)

        hora_final = f"{total_minutos // 60:02d}:{total_minutos % 60:02d}"

        return fecha_final, hora_final

    def _parsear_repeticiones(self, valor: Any) -> int:
        texto = self._limpiar_valor(valor)
        if not texto:
            return 1

        try:
            if isinstance(valor, (int, float)):
                return int(valor)
            if re.fullmatch(r'\d+', texto):
                return int(texto)
            if re.fullmatch(r'\d+[.,]\d+', texto):
                return int(float(texto.replace(',', '.')))
        except Exception as e:
            logger.debug(
                "No se pudo parsear repeticiones: %s | valor=%s",
                e,
                valor,
                exc_info=logger.isEnabledFor(logging.DEBUG),
            )

        return 1

    def _limpiar_valor(self, valor: Any) -> Optional[str]:
        if valor is None:
            return None

        texto = str(valor)
        texto = texto.replace('\r', ' ').replace('\n', ' ').replace('\t', ' ')
        texto = re.sub(r'\s+', ' ', texto).strip()
        if texto.lower() in ('none', 'n/a', 'na', 'sin dato', 'sin data'):
            return None

        return texto

    def _identificar_tipo_encabezado(self, texto: Any) -> Optional[str]:
        """Identifica si un valor corresponde a un encabezado conocido usando coincidencia exacta y alias normalizados."""
        if texto is None:
            return None

        limpio = self._limpiar_valor(texto)
        if not limpio:
            return None

        normalizado = self._normalizar_texto(limpio)
        if not normalizado:
            return None

        for tipo, aliases in self._ALIAS_ENCABEZADOS.items():
            for alias in aliases:
                pattern = re.compile(rf"^(?<![a-z0-9]){re.escape(alias)}(?![a-z0-9])$")
                if pattern.fullmatch(normalizado):
                    return tipo

        if normalizado in {'objetivo supervisor', 'supervisor objetivo'}:
            return 'encabezado_general'

        return None

    def _es_encabezado(self, texto: str) -> bool:
        return self._identificar_tipo_encabezado(texto) is not None

    def _formatear_fecha(self, fecha) -> str:
        """Convierte diversos formatos de fecha a YYYY-MM-DD."""
        if isinstance(fecha, str):
            for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                try:
                    return datetime.strptime(fecha, fmt).strftime('%Y-%m-%d')
                except ValueError:
                    continue
            raise ValueError(f"Formato de fecha no reconocido: {fecha}")

        if isinstance(fecha, (datetime, date)):
            return fecha.strftime('%Y-%m-%d')

        raise ValueError(f"Tipo de fecha no soportado: {type(fecha)}")

    def _formatear_hora(self, hora) -> str:
        fecha_dummy, hora_norm = self._normalizar_hora_y_fecha(
            hora,
            date.today()
        )
        return hora_norm

    def _obtener_supervisor_id(
        self,
        nombre_supervisor: str,
    ) -> Optional[int]:
        """Busca supervisor existente por nombre normalizado."""

        if not nombre_supervisor:
            return None

        nombre = str(nombre_supervisor).strip()

        if not nombre:
            return None

        # Asegurar que los caches estén inicializados
        if not self._cache_inicializado:
            self._inicializar_caches()

        normalized = self._normalizar_texto(nombre)
        return self._cache_supervisores.get(normalized)
    
    def _obtener_objetivo_id(
        self,
        nombre_objetivo: str
    ) -> Optional[int]:
        """Busca objetivo existente por nombre normalizado."""

        if not nombre_objetivo:
            return None

        nombre = str(nombre_objetivo).strip()

        if not nombre:
            return None

        # Asegurar que los caches estén inicializados
        if not self._cache_inicializado:
            self._inicializar_caches()

        normalized = self._normalizar_texto(nombre)
        return self._cache_objetivos.get(normalized)
    
    def _es_duplicado(self, supervisor_id: int, objetivo_id: int, fecha_operativa: str, hora: str, turno: str) -> bool:
        """Verifica si una pasada ya existe (duplicada)."""
        try:
            resultados = gestor_db.ejecutar(
                """
                SELECT 1
                FROM pasadas
                WHERE fecha = ?
                  AND hora = ?
                  AND turno = ?
                  AND supervisor_id = ?
                  AND objetivo_id = ?
                LIMIT 1
                """,
                (fecha_operativa, hora, turno, supervisor_id, objetivo_id),
            )
            return bool(resultados)
        except Exception as e:
            logger.warning("No se pudo verificar duplicado (asumiendo no duplicado): %s", e)
            if logger.isEnabledFor(logging.DEBUG):
                logger.debug("Traceback _es_duplicado", exc_info=True)
            return False

    def _normalizar_texto(self, valor: Any) -> str:
        texto = str(valor).strip().lower()
        texto = unicodedata.normalize('NFKD', texto)
        texto = ''.join(ch for ch in texto if not unicodedata.combining(ch))
        texto = re.sub(r'[^a-z0-9 ]+', ' ', texto)
        texto = re.sub(r'\s+', ' ', texto).strip()
        return texto


_importador: Optional[ImportadorUniversal] = None


def get_importador() -> ImportadorUniversal:
    """Obtiene el importador universal (inicialización diferida)."""
    global _importador
    if _importador is None:
        _importador = ImportadorUniversal()
    return _importador
