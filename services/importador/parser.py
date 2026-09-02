"""
parser.py

Lectura "cruda" (sin normalizar ni validar) del Excel de Control de Recorridos.

Responsabilidades de este módulo, y SOLO estas:
  - Detectar qué hojas son hojas de datos (pasadas) vs. hojas de catálogo/listado
    u otras hojas irrelevantes (vacías, etc.), basándose en el patrón de
    encabezado NO/OBJETIVO/TURNO/MOVIL/HORA/SUPERVISOR — nunca en el nombre
    de la hoja.
  - Extraer, por hoja, todas las PasadaCruda (una por cada fila x cada uno de
    los hasta 3 bloques de columnas), incluyendo las vacías (bloque presente
    pero sin turno/movil/hora/supervisor cargados).
  - Parsear el nombre de la hoja (ej. "1-7 (D)") a fecha + turno (Fase 3).

Todo lo que sea normalización de hora, cálculo de fecha operativa, detección
de duplicados, matching de objetivos/supervisores, etc. es responsabilidad
de etapas posteriores del pipeline de importación, NO de este módulo.
"""

from __future__ import annotations

import re
import logging
from dataclasses import dataclass
from datetime import date
from typing import Any, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .modelos import Problema

# Encabezados esperados, en orden, para cada bloque de columnas.
_ENCABEZADO_BLOQUE = ["NO", "OBJETIVO", "TURNO", "MOVIL", "HORA", "SUPERVISOR"]

# Cuántas filas desde arriba se escanean buscando la fila de encabezado.
# (En los datos reales aparece en la fila 2 o 3 según la hoja, pero se deja
# margen por si algún mes cambia el formato del título superior).
_MAX_FILAS_BUSQUEDA_ENCABEZADO = 15

# Tope de filas de datos a recorrer tras el encabezado, como cinturón de
# seguridad ante hojas corruptas/atípicas que nunca corten el loop.
_MAX_FILAS_DATOS = 500

# Cache del workbook abierto en modo NO read_only (necesario para poder
# indexar hojas por nombre repetidamente sin pagar el costo de carga -que
# puede ser de varios segundos en archivos grandes- en cada llamada a
# leer_pasadas_crudas). Se cachea por path.
_CACHE_WORKBOOKS: dict[str, "openpyxl.Workbook"] = {}
logger = logging.getLogger(__name__)


def _get_workbook(path: str) -> "openpyxl.Workbook":
    wb = _CACHE_WORKBOOKS.get(path)
    if wb is None:
        wb = openpyxl.load_workbook(path, data_only=True)
        _CACHE_WORKBOOKS[path] = wb
    return wb


@dataclass
class PasadaCruda:
    """Una fila cruda de un bloque de columnas, tal cual viene del Excel.

    Ningún campo está validado ni normalizado todavía. `hora` puede ser
    datetime.time, str (con separador ';' u otro), int/float, o None.
    """

    hoja: str
    fila_excel: int  # número de fila real en la hoja (1-indexed), para trazabilidad
    bloque_tabla: int  # 1, 2 o 3
    no: Optional[Any]
    objetivo: Optional[str]
    turno: Optional[str]
    movil: Optional[str]
    hora: Optional[Any]
    supervisor: Optional[str]

    def esta_vacia(self) -> bool:
        """True si el bloque no tiene ningún dato de pasada cargado.

        El NO/OBJETIVO no cuentan para esto: son metadata del objetivo, no
        de la pasada en sí. Una pasada real necesita al menos turno, movil,
        hora o supervisor.
        """
        return not any(
            _valor_no_vacio(v) for v in (self.turno, self.movil, self.hora, self.supervisor)
        )


def _valor_no_vacio(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, str) and v.strip() == "":
        return False
    return True


def _norm_texto(v: Any) -> str:
    """Normaliza un valor de celda de encabezado para comparar: str, strip, upper."""
    if v is None:
        return ""
    return str(v).strip().upper()


def _normalizar_objetivo(v: Any) -> Optional[str]:
    if v is None:
        return None
    texto = re.sub(r"\s+", " ", str(v)).strip()
    return texto or None


def _es_numerico(v: Any) -> bool:
    """True si v representa un número (el NO de una fila de datos)."""
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        try:
            float(v.strip())
            return True
        except ValueError:
            return False
    return False


def _fila_contiene_observaciones(ws: Worksheet, fila: int, max_col: int) -> bool:
    for c in range(1, max_col + 1):
        v = ws.cell(row=fila, column=c).value
        if isinstance(v, str) and "OBSERVACIONES" in v.strip().upper():
            return True
    return False


def _detectar_bloques(ws: Worksheet) -> Optional[tuple[int, list[int]]]:
    """Busca la fila de encabezado y las columnas donde arranca cada bloque.

    Un bloque arranca en la columna `c` de la fila `r` si:
        ws[r][c]   == "NO"
        ws[r][c+1] == "OBJETIVO"
        ws[r][c+2] == "TURNO"
        ws[r][c+3] == "MOVIL"
        ws[r][c+4] == "HORA"
        ws[r][c+5] == "SUPERVISOR"

    Devuelve (fila_encabezado, [columnas_inicio_bloque, ...]) o None si la
    hoja no tiene ninguna fila con ese patrón (p. ej. hoja "Listado" o una
    hoja vacía).
    """
    # ws.max_row / ws.max_column pueden venir None en modo read_only con
    # algunos archivos, así que no se usan como límite superior confiable:
    # se escanea un rango fijo generoso.
    max_row = _MAX_FILAS_BUSQUEDA_ENCABEZADO
    max_col = 60

    for r in range(1, max_row + 1):
        columnas_bloque: list[int] = []
        for c in range(1, max_col + 1):
            if _norm_texto(ws.cell(row=r, column=c).value) != "NO":
                continue
            # candidato: chequear que las siguientes 5 columnas matcheen
            valores = [
                _norm_texto(ws.cell(row=r, column=c + offset).value)
                for offset in range(1, len(_ENCABEZADO_BLOQUE))
            ]
            if valores == _ENCABEZADO_BLOQUE[1:]:
                columnas_bloque.append(c)
        if columnas_bloque:
            return r, columnas_bloque

    return None


def leer_hojas_de_datos(path: str) -> list[str]:
    """Devuelve los nombres de las hojas que son hojas de datos de pasadas.

    Se excluye cualquier hoja que no tenga al menos un bloque de encabezado
    NO/OBJETIVO/TURNO/MOVIL/HORA/SUPERVISOR (p. ej. la hoja "Listado", que
    es catálogo, u hojas vacías/sobrantes). No se usa el nombre de la hoja
    para decidir esto.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        hojas_datos = []
        for nombre in wb.sheetnames:
            if not _PATRON_HOJA_DATOS.fullmatch(nombre.strip()):
                logger.info("Hoja excluida del importador: hoja=%s motivo=nombre fuera del patrón de día", nombre)
                continue
            ws = wb[nombre]
            if _detectar_bloques(ws) is not None:
                hojas_datos.append(nombre)
        return hojas_datos
    finally:
        wb.close()


def leer_pasadas_crudas(path: str, nombre_hoja: str) -> list[PasadaCruda]:
    """Extrae todas las PasadaCruda (vacías y no vacías) de una hoja de datos.

    Recorre las filas de datos (debajo del encabezado) hasta que la columna
    NO del primer bloque deja de ser numérica, o hasta que la fila contiene
    la palabra "OBSERVACIONES" en alguna celda. Genera una PasadaCruda por
    cada uno de los bloques detectados (hasta 3), incluso si ese bloque
    está vacío para esa fila — el filtrado de vacías es responsabilidad del
    llamador.

    Si el OBJETIVO de un bloque viene vacío pero otro bloque de la misma
    fila sí lo tiene, se usa ese nombre (para no perderlo por celdas
    vacías que a veces trae el Excel).
    """
    wb = _get_workbook(path)
    ws = wb[nombre_hoja]
    deteccion = _detectar_bloques(ws)
    if deteccion is None:
        raise ValueError(
            f"La hoja '{nombre_hoja}' no tiene el patrón de encabezado "
            "NO/OBJETIVO/TURNO/MOVIL/HORA/SUPERVISOR; no es una hoja de datos."
        )
    fila_encabezado, columnas_bloque = deteccion
    max_col = max(columnas_bloque) + len(_ENCABEZADO_BLOQUE)

    pasadas: list[PasadaCruda] = []
    fila = fila_encabezado + 1
    max_fila = ws.max_row

    while fila <= max_fila:

        if _fila_contiene_observaciones(ws, fila, max_col):
            logger.info(
                "Fin de tabla: hoja=%s fila=%d motivo=OBSERVACIONES",
                nombre_hoja,
                fila,
            )
            break

        # --- extraer los datos crudos de cada bloque presente en la fila ---
        datos_bloques: list[dict[str, Any]] = []
        for col_no in columnas_bloque:
            col_objetivo = col_no + 1
            col_turno = col_no + 2
            col_movil = col_no + 3
            col_hora = col_no + 4
            col_supervisor = col_no + 5

            datos_bloques.append(
                {
                    "no": ws.cell(row=fila, column=col_no).value,
                    "objetivo": ws.cell(row=fila, column=col_objetivo).value,
                    "turno": ws.cell(row=fila, column=col_turno).value,
                    "movil": ws.cell(row=fila, column=col_movil).value,
                    "hora": ws.cell(row=fila, column=col_hora).value,
                    "supervisor": ws.cell(row=fila, column=col_supervisor).value,
                }
            )

        # El primer bloque puede estar vacío aunque otro bloque de la misma
        # fila tenga una pasada. Solo se termina la lectura cuando no hay
        # datos en ningún bloque, evitando perder pasadas de los bloques 2/3.
        if not any(
            _valor_no_vacio(d[campo])
            for d in datos_bloques
            for campo in ("no", "objetivo", "turno", "movil", "hora", "supervisor")
        ):
            logger.info(
                "Fin de tabla: hoja=%s fila=%d motivo=fila sin datos en ningún bloque",
                nombre_hoja,
                fila,
            )
            break

        # --- rellenar objetivo faltante desde otro bloque de la misma fila ---
        objetivo_comun = None
        for d in datos_bloques:
            if _valor_no_vacio(d["objetivo"]):
                objetivo_comun = _normalizar_objetivo(d["objetivo"])
                break

        for idx, d in enumerate(datos_bloques, start=1):
            objetivo = (
                _normalizar_objetivo(d["objetivo"])
                if _valor_no_vacio(d["objetivo"])
                else objetivo_comun
            )
            pasadas.append(
                PasadaCruda(
                    hoja=nombre_hoja,
                    fila_excel=fila,
                    bloque_tabla=idx,
                    no=d["no"],
                    objetivo=objetivo,
                    turno=d["turno"],
                    movil=d["movil"],
                    hora=d["hora"],
                    supervisor=d["supervisor"],
                )
            )

        fila += 1

    return pasadas


# ---------------------------------------------------------------------------
# FASE 3 — Parseo del nombre de hoja a fecha operativa + turno
# ---------------------------------------------------------------------------

# Nombres de hoja observados en archivos reales: "1-7 (D)", "1-7 (N)",
# "6-7(D)" (sin espacio antes del paréntesis), "31-7 (N)".
# Grupo 1: día, grupo 2: mes, grupo 3: letra/palabra de turno.
_PATRON_NOMBRE_HOJA = re.compile(
    r"^\s*(\d{1,2})\s*-\s*(\d{1,2})\s*\(?\s*([A-ZÁÉÍÓÚa-záéíóú]+)\s*\)?\s*$"
)
_PATRON_HOJA_DATOS = re.compile(r"^\d{1,2}-\d{1,2} \((D|N)\)$")

# Variantes de turno aceptadas para el nombre de hoja (punto 4 de la
# especificación). Se mantiene local a este módulo, sin depender todavía
# de normalizador.normalizar_turno (que se implementa recién en la Fase 5
# y opera sobre la celda TURNO de cada pasada, no sobre el nombre de hoja).
_VARIANTES_DIURNO = {"D", "DIA", "DÍA", "DIURNO"}
_VARIANTES_NOCTURNO = {"N", "NOCHE", "NOCTURNO"}


def _normalizar_turno_de_nombre_hoja(texto: str) -> Optional[str]:
    texto_norm = texto.strip().upper()
    if texto_norm in _VARIANTES_DIURNO:
        return "D"
    if texto_norm in _VARIANTES_NOCTURNO:
        return "N"
    return None


def parsear_nombre_hoja(
    nombre_hoja: str, anio: int
) -> tuple[Optional[date], Optional[str], Optional[Problema]]:
    """Parsea el nombre de una hoja (ej. "1-7 (D)", "6-7(N)") a fecha y turno.

    El año no viene en el nombre de hoja, así que se recibe como parámetro
    configurable en esta etapa de prueba (punto 3 de la especificación).

    Si el nombre no matchea el patrón esperado, o el turno indicado no es
    una variante reconocida, o la fecha resultante no es válida (ej. día 32),
    NO se adivina ni se descarta silenciosamente: se devuelve un Problema de
    tipo "error_critico" con el detalle, para que quede visible en el
    análisis (punto 6 de la especificación, aplicado también acá).

    Devuelve (fecha, turno, None) si el parseo fue exitoso, o
    (None, None, Problema) si falló.
    """
    match = _PATRON_NOMBRE_HOJA.match(nombre_hoja or "")

    if not match:
        return None, None, Problema(
            tipo="error_critico",
            descripcion=(
                f"No se pudo interpretar el nombre de la hoja '{nombre_hoja}'. "
                "Se esperaba un formato como '1-7 (D)' o '6-7(N)' "
                "(día-mes seguido del turno)."
            ),
            hoja=nombre_hoja,
        )

    dia_texto, mes_texto, turno_texto = match.groups()
    dia, mes = int(dia_texto), int(mes_texto)

    turno = _normalizar_turno_de_nombre_hoja(turno_texto)
    if turno is None:
        return None, None, Problema(
            tipo="error_critico",
            descripcion=(
                f"El turno '{turno_texto}' indicado en el nombre de la hoja "
                f"'{nombre_hoja}' no es una variante reconocida de Diurno/Nocturno."
            ),
            hoja=nombre_hoja,
            valor_problema=turno_texto,
        )

    try:
        fecha = date(anio, mes, dia)
    except ValueError:
        return None, None, Problema(
            tipo="error_critico",
            descripcion=(
                f"La fecha derivada del nombre de la hoja '{nombre_hoja}' "
                f"(día={dia}, mes={mes}, año={anio}) no es una fecha válida."
            ),
            hoja=nombre_hoja,
            valor_problema=nombre_hoja,
        )

    return fecha, turno, None


# ---------------------------------------------------------------------------
# Prueba manual contra el archivo real
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys

    ruta = sys.argv[1] if len(sys.argv) > 1 else "CONTROL_RECORRIDOS_JULIO_2026.xlsx"

    hojas = leer_hojas_de_datos(ruta)
    print(f"Hojas de datos detectadas: {len(hojas)}")
    print(hojas)
    print()

    total_no_vacias = 0
    resumen = []
    for hoja in hojas:
        pasadas = leer_pasadas_crudas(ruta, hoja)
        no_vacias = [p for p in pasadas if not p.esta_vacia()]
        total_no_vacias += len(no_vacias)
        resumen.append((hoja, len(pasadas), len(no_vacias)))
        print(f"{hoja:15s} -> total bloques: {len(pasadas):4d}  |  no vacías: {len(no_vacias):3d}")

    print()
    print(f"TOTAL pasadas no vacías en todo el archivo: {total_no_vacias}")

    print()
    print("--- Parseo de nombres de hoja (Fase 3) ---")
    for hoja in hojas:
        fecha, turno, problema = parsear_nombre_hoja(hoja, anio=2026)
        if problema:
            print(f"  {hoja!r:15} -> ERROR: {problema.descripcion}")
        else:
            print(f"  {hoja!r:15} -> fecha={fecha}  turno={turno}")