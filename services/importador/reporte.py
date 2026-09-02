"""
reporte.py

Fase 10: motor de análisis completo.

`analizar_excel()` orquesta todo el pipeline de importación en modo
SOLO LECTURA (no escribe nada en la base):

    parser (Fase 1-3)
        -> normalizador (Fase 4-5)
        -> matcher (Fase 6-7)
        -> duplicados (Fase 8)
        -> validador (Fase 9)
        -> ResultadoAnalisis (resumen)

`generar_reporte_detallado()` vuelca cada Problema detectado a un Excel
(.xlsx) para el botón "Descargar análisis detallado". Se eligió Excel en
vez de PDF: el contenido es tabular (una fila por Problema, con columnas
tipo/hoja/fila/objetivo/valor/descripción) que alguien va a querer
filtrar y ordenar, y el pipeline ya depende de openpyxl para todo lo
demás, así que no se suma ninguna dependencia nueva.
"""

from __future__ import annotations

import io
import logging
from typing import Optional

from . import duplicados, matcher, parser
from .modelos import PasadaNormalizada, Problema, ResultadoAnalisis
from .normalizador import (
    determinar_fecha_operativa_y_calendario,
    normalizar_hora,
    normalizar_turno,
    resolver_turno_con_prioridad,
)
from .validador import validar

import openpyxl
from openpyxl.styles import Font


logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Fase 1-5: de PasadaCruda a PasadaNormalizada, generando los Problema que
# faltaba cerrar (hora inválida / hora normalizada), señalados en el
# análisis previo de validador.py.
# ---------------------------------------------------------------------------


def _construir_pasadas_normalizadas(
    path: str, anio: int
) -> tuple[list[PasadaNormalizada], list[Problema], int, int]:
    """Devuelve (pasadas_normalizadas, problemas, hojas_encontradas,
    pasadas_detectadas).

    `pasadas_detectadas` cuenta todos los bloques no vacíos. Las filas sin
    hora o con hora inválida quedan registradas como advertencias y no se
    convierten en PasadaNormalizada.
    """
    hojas = parser.leer_hojas_de_datos(path)
    problemas: list[Problema] = []
    pasadas_normalizadas: list[PasadaNormalizada] = []
    pasadas_detectadas = 0

    for hoja in hojas:
        fecha_hoja, turno_hoja, problema_hoja = parser.parsear_nombre_hoja(hoja, anio)
        if problema_hoja is not None:
            # No se puede fechar/turnar ninguna pasada de esta hoja: se
            # reporta el error y se saltea la hoja entera.
            problemas.append(problema_hoja)
            logger.error(
                "Hoja descartada: hoja=%s fila=- bloque=- motivo=%s",
                hoja,
                problema_hoja.descripcion,
            )
            continue

        crudas = parser.leer_pasadas_crudas(path, hoja)
        crudas_no_vacias = [c for c in crudas if not c.esta_vacia()]
        pasadas_detectadas += len(crudas_no_vacias)
        for cruda in crudas_no_vacias:
            # --- Fase 4: hora ---------------------------------------------
            resultado_hora = normalizar_hora(cruda.hora)

            if resultado_hora.error is not None:
                if resultado_hora.hora is not None and "hora incompleta" in resultado_hora.error:
                    problemas.append(
                        Problema(
                            tipo="advertencia",
                            descripcion=resultado_hora.error,
                            hoja=hoja,
                            objetivo=cruda.objetivo,
                            valor_problema=cruda.hora,
                            fila_excel=cruda.fila_excel,
                        )
                    )
                else:
                    logger.warning(
                        "Pasada descartada: hoja=%s fila=%d bloque=%d objetivo=%r motivo=hora inválida: %s",
                        hoja, cruda.fila_excel, cruda.bloque_tabla, cruda.objetivo,
                        resultado_hora.error,
                    )
                    problemas.append(
                        Problema(
                            tipo="advertencia",
                            descripcion=(
                                f"Pasada omitida: {resultado_hora.error}. "
                                "El supervisor no cargó una hora válida; la pasada "
                                "queda visible en este informe, pero no se importa."
                            ),
                            hoja=hoja,
                            objetivo=cruda.objetivo,
                            valor_problema=cruda.hora,
                            fila_excel=cruda.fila_excel,
                        )
                    )
                    continue

            if resultado_hora.hora is None:
                logger.warning(
                    "Pasada retenida como incompleta: hoja=%s fila=%d bloque=%d objetivo=%r motivo=hora ausente",
                    hoja, cruda.fila_excel, cruda.bloque_tabla, cruda.objetivo,
                )
                # Celda de hora vacía, pero la fila no está vacía en su
                # conjunto (tiene movil/turno/supervisor cargado). No es
                # el mismo caso que "hora inválida" (normalizar_hora no
                # marca esto como error porque una celda vacía es válida
                # cuando toda la fila está vacía) — acá sí importa,
                # porque hora es obligatoria en PasadaNormalizada.
                problemas.append(
                    Problema(
                        tipo="advertencia",
                        descripcion=(
                            "Pasada omitida: el supervisor no cargó la hora. "
                            "La pasada queda visible en este informe, pero "
                            "no se importa."
                        ),
                        hoja=hoja,
                        objetivo=cruda.objetivo,
                        fila_excel=cruda.fila_excel,
                    )
                )
                continue

            if resultado_hora.fue_normalizada:
                problemas.append(
                    Problema(
                        tipo="advertencia",
                        descripcion=(
                            f"Hora normalizada automáticamente desde "
                            f"'{cruda.hora}' a {resultado_hora.hora.strftime('%H:%M')}."
                        ),
                        hoja=hoja,
                        objetivo=cruda.objetivo,
                        valor_problema=cruda.hora,
                        fila_excel=cruda.fila_excel,
                    )
                )


            # --- Fase 5: turno ----------------------------------------------
            if cruda.turno is not None and str(cruda.turno).strip() != "":
                resultado_turno = normalizar_turno(
                    cruda.turno, hoja=hoja, objetivo=cruda.objetivo, fila_excel=cruda.fila_excel
                )
                if resultado_turno.problema is not None:
                    problemas.append(resultado_turno.problema)
                turno_celda = resultado_turno.turno
            else:
                turno_celda = None

            turno_final, problema_prioridad = resolver_turno_con_prioridad(
                turno_celda,
                turno_hoja,
                hoja=hoja,
                objetivo=cruda.objetivo,
                fila_excel=cruda.fila_excel,
            )
            if problema_prioridad is not None:
                problemas.append(problema_prioridad)

            if turno_final is None:
                # Turno de celda inválido Y sin turno de hoja como
                # respaldo (no debería pasar, ya que parsear_nombre_hoja
                # ya garantizó turno_hoja válido, pero se cubre el caso
                # límite igual en vez de asumir).
                logger.warning(
                    "Pasada descartada: hoja=%s fila=%d bloque=%d objetivo=%r motivo=turno no resoluble",
                    hoja, cruda.fila_excel, cruda.bloque_tabla, cruda.objetivo,
                )
                continue

            fecha_operativa, fecha_calendario = determinar_fecha_operativa_y_calendario(
                fecha_hoja, turno_final, resultado_hora.hora
            )

            pasadas_normalizadas.append(
                PasadaNormalizada(
                    hoja=hoja,
                    fila_excel=cruda.fila_excel,
                    bloque_tabla=cruda.bloque_tabla,
                    fecha_operativa=fecha_operativa,
                    fecha_calendario=fecha_calendario,
                    turno=turno_final,
                    turno_hoja=turno_hoja,
                    hora=resultado_hora.hora,
                    movil=cruda.movil,
                    objetivo_nombre=(cruda.objetivo or "").strip() if cruda.objetivo else "",
                    supervisor_nombre=cruda.supervisor,
                )
            )
    return pasadas_normalizadas, problemas, len(hojas), pasadas_detectadas


# ---------------------------------------------------------------------------
# Orquestador principal
# ---------------------------------------------------------------------------


def analizar_excel(
    path: str,
    anio: int,
    conexion_bd,
    forzar_sobrescritura: bool = False,
) -> ResultadoAnalisis:
    """Analiza un Excel de Control de Recorridos de punta a punta, SIN
    escribir nada en la base. Devuelve un ResultadoAnalisis listo para
    mostrar en la pantalla de revisión."""

    # --- Fase 1-5: lectura + normalización -----------------------------
    pasadas, problemas, hojas_encontradas, pasadas_detectadas = _construir_pasadas_normalizadas(
        path, anio
    )

    # --- Fase 6-7: matching ----------------------------------------------
    objetivos_bd = matcher.obtener_objetivos_bd(conexion_bd)
    supervisores_bd = matcher.obtener_supervisores_bd(conexion_bd)

    resultados_match_objetivo = matcher.aplicar_matching_objetivos(pasadas, objetivos_bd)
    resultados_match_supervisor = matcher.aplicar_matching_supervisores(pasadas, supervisores_bd)

    objetivos_para_revisar = sum(
        1 for r in resultados_match_objetivo if r.tipo != "exacto"
    )
    supervisores_para_revisar = sum(
        1 for r in resultados_match_supervisor if r.tipo != "exacto"
    )

    # --- Fase 8: duplicados internos + existencia en base -----------------
    pasadas_tras_dedupe, descartadas = duplicados.detectar_duplicados_internos(pasadas)
    nuevas, existentes = duplicados.detectar_pasadas_existentes(
        pasadas_tras_dedupe, conexion_bd, forzar_sobrescritura=forzar_sobrescritura
    )

        # --- Fase 6-7 (cont.): volcar matching no exacto como Problema -------
    def _problemas_de_matching(resultados, atributo_nombre, nombre_candidato):
        extra: list[Problema] = []
        for resultado in resultados:
            if resultado.tipo == "exacto":
                continue
            clave = matcher.normalizar_nombre(resultado.nombre_excel)
            afectadas = [
                p for p in pasadas_tras_dedupe
                if getattr(p, atributo_nombre)
                and matcher.normalizar_nombre(getattr(p, atributo_nombre)) == clave
            ]
            if resultado.tipo == "sugerencias":
                nombres = ", ".join(nombre_candidato(s) for s in resultado.sugerencias[:3])
                detalle = f"Sin match exacto. Sugerencias: {nombres}."
            else:
                detalle = "Sin match exacto ni sugerencias cercanas."
            destinos = afectadas or [None]  # no perder el problema si no hay pasada asociada
            for p in destinos:
                logger.warning(
                    "Matching pendiente: hoja=%s fila=%s bloque=%s objetivo=%r campo=%s motivo=sin coincidencia exacta",
                    p.hoja if p else "-",
                    p.fila_excel if p else "-",
                    p.bloque_tabla if p else "-",
                    resultado.nombre_excel,
                    atributo_nombre,
                )
                extra.append(
                    Problema(
                        tipo="para_revisar",
                        descripcion=f"'{resultado.nombre_excel}' no reconocido. {detalle}",
                        hoja=p.hoja if p else None,
                        objetivo=resultado.nombre_excel,
                        fila_excel=p.fila_excel if p else None,
                        valor_problema=resultado,  # <-- AGREGAR ESTA LÍNEA
                    )
                )
        return extra

    problemas += _problemas_de_matching(
        resultados_match_objetivo, "objetivo_nombre", lambda s: s.objetivo.nombre
    )
    problemas += _problemas_de_matching(
        resultados_match_supervisor, "supervisor_nombre", lambda s: s.supervisor.nombre
    )

    pasadas_actualizar = sum(1 for p in existentes if p.accion == "actualizar")
    pasadas_omitir = sum(1 for p in existentes if p.accion == "omitir")
    # Las de objetivo_id None también terminan en `nuevas` (ver
    # duplicados.detectar_pasadas_existentes), con accion=None: no cuentan
    # como "nueva" confirmada todavía porque su matching sigue pendiente,
    # pero sí como pasada final del análisis.
    pasadas_nuevas = sum(1 for p in nuevas if p.accion == "nueva")

    # --- Fase 9: validación de negocio ------------------------------------
    problemas += validar(pasadas_tras_dedupe, contexto=None)

    errores_criticos = sum(1 for p in problemas if p.tipo == "error_critico")
    advertencias = sum(1 for p in problemas if p.tipo == "advertencia")
    pasadas_sin_hora = sum("no cargó la hora" in p.descripcion for p in problemas)

    return ResultadoAnalisis(
        pasadas=pasadas_tras_dedupe,
        problemas=problemas,
        total_pasadas=len(pasadas_tras_dedupe),
        pasadas_nuevas=pasadas_nuevas,
        pasadas_actualizar=pasadas_actualizar,
        pasadas_omitir=pasadas_omitir,
        objetivos_para_revisar=objetivos_para_revisar,
        supervisores_para_revisar=supervisores_para_revisar,
        errores_criticos=errores_criticos,
        advertencias=advertencias,
        hojas_encontradas=hojas_encontradas,
        pasadas_detectadas=pasadas_detectadas,
        pasadas_sin_hora=pasadas_sin_hora,
        pasadas_duplicadas=len(descartadas),
    )


def generar_resumen_texto(resultado: ResultadoAnalisis) -> str:
    """Arma el bloque de texto pedido para mostrar el resumen del análisis."""
    return (
        "ANÁLISIS DEL ARCHIVO\n"
        f"Hojas encontradas: {resultado.hojas_encontradas}\n"
        f"Pasadas detectadas: {resultado.pasadas_detectadas}\n"
        f"Pasadas sin hora: {resultado.pasadas_sin_hora}\n"
        f"Pasadas nuevas: {resultado.pasadas_nuevas}\n"
        f"Pasadas ya existentes: {resultado.pasadas_omitir + resultado.pasadas_actualizar}\n"
        f"Duplicados descartados: {resultado.pasadas_duplicadas}\n"
        f"Objetivos no reconocidos: {resultado.objetivos_para_revisar}\n"
        f"Supervisores no reconocidos: {resultado.supervisores_para_revisar}\n"
        f"Errores críticos: {resultado.errores_criticos}\n"
        f"Advertencias: {resultado.advertencias}\n"
    )


# ---------------------------------------------------------------------------
# Reporte detallado descargable
# ---------------------------------------------------------------------------


def generar_reporte_detallado(resultado: ResultadoAnalisis) -> bytes:
    """Genera un .xlsx con el detalle completo de cada Problema, más una
    hoja de resumen con los mismos números de generar_resumen_texto()."""
    wb = openpyxl.Workbook()

    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    encabezado_font = Font(bold=True)

    filas_resumen = [
        ("Hojas encontradas", resultado.hojas_encontradas),
        ("Pasadas detectadas", resultado.pasadas_detectadas),
        ("Pasadas sin hora", resultado.pasadas_sin_hora),
        ("Pasadas nuevas", resultado.pasadas_nuevas),
        ("Pasadas ya existentes", resultado.pasadas_omitir + resultado.pasadas_actualizar),
        ("Duplicados descartados", resultado.pasadas_duplicadas),
        ("Objetivos no reconocidos", resultado.objetivos_para_revisar),
        ("Supervisores no reconocidos", resultado.supervisores_para_revisar),
        ("Errores críticos", resultado.errores_criticos),
        ("Advertencias", resultado.advertencias),
    ]
    ws_resumen.append(["Análisis del archivo", ""])
    ws_resumen["A1"].font = Font(bold=True, size=13)
    ws_resumen.append(["", ""])
    for etiqueta, valor in filas_resumen:
        ws_resumen.append([etiqueta, valor])
    ws_resumen.column_dimensions["A"].width = 28
    ws_resumen.column_dimensions["B"].width = 12

    ws_problemas = wb.create_sheet("Problemas")
    columnas = ["Tipo", "Hoja", "Fila Excel", "Objetivo", "Valor", "Descripción"]
    ws_problemas.append(columnas)
    for celda in ws_problemas[1]:
        celda.font = encabezado_font

    for p in resultado.problemas:
        ws_problemas.append(
            [
                p.tipo,
                p.hoja or "",
                p.fila_excel if p.fila_excel is not None else "",
                p.objetivo or "",
                str(p.valor_problema) if p.valor_problema is not None else "",
                p.descripcion,
            ]
        )

    ws_horas = wb.create_sheet("Horas normalizadas")
    ws_horas.append([
        "Hoja", "Fila Excel", "Objetivo", "Hora original", "Hora normalizada"
    ])
    for celda in ws_horas[1]:
        celda.font = encabezado_font
    for p in resultado.problemas:
        if "Hora normalizada automáticamente" not in p.descripcion:
            continue
        texto = p.descripcion
        hora_original = texto.split("desde '", 1)[1].split("'", 1)[0]
        hora_normalizada = texto.rsplit(" a ", 1)[-1].rstrip(".")
        ws_horas.append([
            p.hoja or "",
            p.fila_excel if p.fila_excel is not None else "",
            p.objetivo or "",
            hora_original,
            hora_normalizada,
        ])
    for col, ancho in zip("ABCDE", [18, 12, 28, 18, 20]):
        ws_horas.column_dimensions[col].width = ancho

    anchos = [16, 14, 10, 24, 20, 70]
    for col, ancho in zip("ABCDEF", anchos):
        ws_problemas.column_dimensions[col].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()