# =============================================================================
# VESP Organizations - Sistema de Control de Objetivos
# Módulo de exportación de reportes a Excel y PDF
# =============================================================================

import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from services.reportes import generar_reporte_mensual, clasificar_cumplimiento


MESES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]


def _texto_filtro(valor: str | None) -> str:
    return valor or "Todos"


def exportar_pasadas_excel(datos: list, ruta: str, filtros: dict) -> None:
    """Exporta exactamente las pasadas ya filtradas en pantalla."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pasadas"

    ws.merge_cells("A1:D1")
    ws["A1"] = "V.E.S.P Organizations - Pasadas filtradas"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:D2")
    ws["A2"] = (
        f"Fecha: {_texto_filtro(filtros.get('fecha'))} | "
        f"Supervisor: {_texto_filtro(filtros.get('supervisor'))} | "
        f"Turno: {_texto_filtro(filtros.get('turno'))} | "
        f"Búsqueda: {_texto_filtro(filtros.get('busqueda'))}"
    )
    ws["A2"].alignment = Alignment(horizontal="left")

    encabezados = ["Hora", "Turno", "Objetivo", "Supervisor"]
    for col, encabezado in enumerate(encabezados, 1):
        celda = ws.cell(row=4, column=col, value=encabezado)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(fill_type="solid", fgColor="1B5E20")
        celda.alignment = Alignment(horizontal="center")

    for fila, pasada in enumerate(datos, 5):
        for col, valor in enumerate(pasada[1:5], 1):
            ws.cell(row=fila, column=col, value=valor or "")

    for columna, ancho in {"A": 14, "B": 14, "C": 34, "D": 28}.items():
        ws.column_dimensions[columna].width = ancho
    wb.save(ruta)


def exportar_pasadas_pdf(datos: list, ruta: str, filtros: dict) -> None:
    """Exporta exactamente las pasadas ya filtradas en pantalla."""
    doc = SimpleDocTemplate(
        ruta, pagesize=A4,
        rightMargin=1.5 * cm, leftMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm
    )
    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph("<b>V.E.S.P Organizations - Pasadas filtradas</b>", estilos["Title"]),
        Paragraph(
            f"Fecha: {_texto_filtro(filtros.get('fecha'))} | "
            f"Supervisor: {_texto_filtro(filtros.get('supervisor'))} | "
            f"Turno: {_texto_filtro(filtros.get('turno'))} | "
            f"Búsqueda: {_texto_filtro(filtros.get('busqueda'))}",
            estilos["Normal"]
        ),
        Spacer(1, 0.4 * cm),
    ]
    filas = [["Hora", "Turno", "Objetivo", "Supervisor"]]
    filas.extend([[str(valor or "") for valor in pasada[1:5]] for pasada in datos])
    tabla = Table(filas, colWidths=[2.5 * cm, 3 * cm, 7.5 * cm, 5 * cm])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B5E20")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elementos.append(tabla)
    doc.build(elementos)


def _obtener_datos_reporte(anio: int, mes: int, reporte: dict | None = None) -> list:
    """Obtiene el resumen de reporte mensual reutilizando la lógica central."""
    reporte = reporte or generar_reporte_mensual(anio, mes)
    return [
        (
            objetivo['nombre'],
            objetivo['dias_con_pasada'],
            objetivo['dias_sin_pasada'],
            objetivo['cumplimiento_porcentaje']
        )
        for objetivo in reporte['objetivos']
    ]


def exportar_excel(anio: int, mes: int, ruta: str, reporte: dict | None = None, filtros: dict | None = None) -> None:
    """Genera Excel con reporte mensual."""
    # r = (nombre, dias_con_pasada, dias_sin_pasada, cumplimiento_porcentaje)
    resultados = _obtener_datos_reporte(anio, mes, reporte)
    filtros = filtros or {}

    wb = Workbook()
    ws = wb.active
    ws.title = f"Reporte {MESES[mes-1]} {anio}"

    try:
        from services.assets import ruta_asset
        img = XLImage(ruta_asset("assets/vesp.png"))
        img.width = 80
        img.height = 80
        ws.add_image(img, "A1")
    except Exception:
        pass

    ws.merge_cells("B1:F2")
    ws["B1"] = "V.E.S.P Organizations - Seguridad Privada"
    ws["B1"].font = Font(bold=True, size=14, color="2E7D32")
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:F3")
    ws["A3"] = f"Reporte mensual - {MESES[mes-1]} {anio}"
    ws["A3"].font = Font(bold=True, size=12)
    ws["A3"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A5:F5")
    ws["A5"] = "Filtros: " + " | ".join(f"{clave}: {valor or 'Todos'}" for clave, valor in filtros.items())

    ws.merge_cells("A4:F4")
    ws["A4"] = f"Generado el {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A4"].font = Font(size=9, color="888888")
    ws["A4"].alignment = Alignment(horizontal="center")

    encabezados = ["Objetivo", "Días con pasada", "Días sin pasada", "Cumplimiento", "Estado"]
    for col, enc in enumerate(encabezados, 1):
        celda = ws.cell(row=7, column=col, value=enc)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(fill_type="solid", fgColor="1B5E20")
        celda.alignment = Alignment(horizontal="center")

    for fila, r in enumerate(resultados, 8):
        estado, categoria = clasificar_cumplimiento(r[3])
        valores = [r[0], r[1], r[2], f"{r[3]:.1f}%", estado]
        for col, val in enumerate(valores, 1):
            celda = ws.cell(row=fila, column=col, value=val)
            color = {"verde": "C8E6C9", "amarillo": "FFF3CD", "rojo": "FFCDD2"}[categoria]
            celda.fill = PatternFill(fill_type="solid", fgColor=color)
            celda.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 60

    wb.save(ruta)


def exportar_pdf(anio: int, mes: int, ruta: str, reporte: dict | None = None, filtros: dict | None = None) -> None:
    """Genera PDF con reporte mensual."""
    # r = (nombre, dias_con_pasada, dias_sin_pasada, cumplimiento_porcentaje)
    resultados = _obtener_datos_reporte(anio, mes, reporte)
    filtros = filtros or {}

    doc = SimpleDocTemplate(
        ruta, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    estilos = getSampleStyleSheet()
    elementos = []

    try:
        from services.assets import ruta_asset
        logo = RLImage(ruta_asset("assets/vesp.png"), width=2.5*cm, height=2.5*cm)
        datos_header = [[
            logo,
            Paragraph(
                "<b><font color='#2E7D32' size=14>V.E.S.P Organizations</font></b>"
                "<br/><font size=10>Seguridad Privada</font>",
                estilos["Normal"]
            ),
            Paragraph(
                f"<font size=9 color='grey'>Generado el<br/>"
                f"{datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}</font>",
                estilos["Normal"]
            )
        ]]
        tabla_header = Table(datos_header, colWidths=[3*cm, 10*cm, 4*cm])
        tabla_header.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (2, 0), (2, 0), "RIGHT"),
        ]))
        elementos.append(tabla_header)
    except Exception:
        elementos.append(Paragraph("<b>V.E.S.P Organizations</b>", estilos["Title"]))

    elementos.append(Spacer(1, 0.5*cm))
    elementos.append(Paragraph(
        "Filtros: " + " | ".join(f"{clave}: {valor or 'Todos'}" for clave, valor in filtros.items()),
        estilos["Normal"]
    ))
    elementos.append(Paragraph(
        f"<b>Reporte mensual - {MESES[mes-1]} {anio}</b>",
        estilos["Title"]
    ))
    elementos.append(Spacer(1, 0.5*cm))

    datos = [["Objetivo", "Días con pasada", "Días sin pasada", "Cumplimiento", "Estado"]]
    for r in resultados:
        estado, _ = clasificar_cumplimiento(r[3])
        datos.append([r[0], str(r[1]), str(r[2]), f"{r[3]:.1f}%", estado])

    tabla = Table(datos, colWidths=[5.5*cm, 3*cm, 3*cm, 2.5*cm, 3*cm])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B5E20")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWHEIGHT", (0, 0), (-1, -1), 20),
    ]))

    for i, r in enumerate(resultados, 1):
        _, categoria = clasificar_cumplimiento(r[3])
        color = colors.HexColor({"verde": "#C8E6C9", "amarillo": "#FFF3CD", "rojo": "#FFCDD2"}[categoria])
        tabla.setStyle(TableStyle([("BACKGROUND", (0, i), (-1, i), color)]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 0.5*cm))

    total = len(resultados)
    cumplen = sum(1 for r in resultados if clasificar_cumplimiento(r[3])[0] == "Cumplió")
    elementos.append(Paragraph(
        f"<b>Resumen:</b> {cumplen} de {total} objetivos cumplen el 50% o más de cobertura.",
        estilos["Normal"]
    ))

    doc.build(elementos)